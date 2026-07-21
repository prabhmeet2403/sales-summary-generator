"""
column_autofit.py
====================
Shared column-width auto-fit logic used by BOTH sheet_copy.py (Sheet 3,
a verbatim copy of the Master workbook's own sheet) and
summary_writer.py (Worksheets 1 and 2, built natively) -- so all three
of the generated workbook's sheets get the SAME sizing behavior,
replacing what used to be an inconsistency: Sheet 3 sized its columns
to content while Sheets 1/2 used fixed, hardcoded widths regardless of
what was actually in them.

The core algorithm is the same for every sheet: measure the widest
VISIBLE text actually in each column (header and data, with a formula
cell measured by its computed result rather than its raw formula
text -- callers resolve that themselves and hand this module the
value to measure), then set that column's width to fit it, bounded to
a sensible [min, max] range so narrow columns don't collapse and one
long comment doesn't blow out the whole sheet.

Row heights are never touched here or anywhere in this pipeline -- every
row keeps whatever height it already had (Excel's own default for
Worksheets 1/2, or whatever height was copied verbatim from the source
for Sheet 3), so every row is the same, uniform height. This is only
possible because the Comments column doesn't use `wrap_text` (see
summary_writer.py) -- Excel automatically expands a row's height to
fit a wrapped cell's full text whenever that row's height is left
unset, so wrapping and uniform row height are mutually exclusive.
Instead, the Comments column is simply allowed a wider maximum than
other columns (see `autofit_worksheet_columns`'s `column_max_width`
parameter) so more of a comment's text fits before it visually
overflows into the empty cells beyond -- the Comments column is always
the last column on these sheets, so nothing else is ever obscured.
"""
from __future__ import annotations

import datetime
from typing import Callable, Dict, Optional

from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# A modest floor keeps genuinely empty/near-empty columns from
# collapsing to an unreadably thin sliver; a ceiling keeps a numeric or
# label column from stretching absurdly wide.
MIN_COLUMN_WIDTH = 6.0
MAX_COLUMN_WIDTH = 50.0

# The Comments column is allowed a much wider maximum than other
# columns -- it's expected to hold long free-text notes, and giving it
# more horizontal room means wrapped text needs fewer lines, which is
# what keeps every row the same height instead of growing to fit a
# long comment.
MAX_COMMENTS_COLUMN_WIDTH = 120.0


def display_text(value: object, number_format: Optional[str]) -> str:
    """Approximate what Excel would actually SHOW for one cell, given
    its value and number format -- used only to measure text length
    for `autofit_worksheet_columns`, never written anywhere. Covers the
    number formats actually observed in this app's output (currency,
    accounting, thousands-separated plain numbers, percentages, dates,
    text, General) with a plain `str(value)` fallback for anything else
    -- a full general-purpose number-format renderer is out of scope
    for a width estimate, and openpyxl has no built-in AutoFit to defer
    to.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%b-%y")
    if isinstance(value, (int, float)):
        nf = number_format or "General"
        if nf in ("General", "@"):
            if isinstance(value, float) and value == int(value):
                return str(int(value))
            return str(value)
        if "$" in nf:
            if value < 0:
                return f"(${abs(value):,.0f})" if "(" in nf else f"-${abs(value):,.0f}"
            return f"${value:,.0f}"
        if "%" in nf:
            return f"{value * 100:.0f}%"
        if "#,##0" in nf:
            return f"{value:,.0f}"
        return str(value)
    return str(value)


# Bold text renders measurably wider, per character, than regular
# weight at the same point size -- a plain character-count estimate
# systematically understates a bold cell's true rendered width. Every
# header, banner, subtotal, and TOTAL row in this workbook is bold
# (see summary_writer.py), so whenever one of those short, bold labels
# (e.g. "Margin", a 3-letter month abbreviation) happens to be the
# widest thing in its column -- typically a column whose actual data
# values are comparatively short too, leaving no other content to mask
# the shortfall -- the plain-text estimate can land the column just
# narrow enough that Excel visually clips the header until the cell is
# selected or the column is manually widened. This is a well-established
# approximation for sans-serif bold-vs-regular character width, not an
# exact metric (a byte-perfect one would require actually rendering the
# font, which openpyxl cannot do) -- applied only to bold cells, so
# regular-weight content (the vast majority of the sheet) is completely
# unaffected.
#
# A pure percentage factor barely helps very short bold labels (15% of
# 3 characters is well under half a character) -- exactly the case
# that clips hardest, since there's so little length to begin with.
# `BOLD_MIN_BONUS` adds a small flat floor on top of the percentage
# scaling so short bold text (month abbreviations, "Total", "Margin")
# gets meaningful help too, while longer bold text still scales
# proportionally rather than by an ever-larger flat amount.
BOLD_WIDTH_FACTOR = 1.15
BOLD_MIN_BONUS = 1.5


def measured_length(cell: Cell, value: object) -> float:
    """The effective character length to use for `cell` when comparing
    column widths -- `display_text`'s plain length, inflated slightly
    for a bold cell (see `BOLD_WIDTH_FACTOR`/`BOLD_MIN_BONUS`) to
    reflect that a bold cell's true rendered width is measurably more
    than its character count alone suggests."""
    text_len = len(display_text(value, cell.number_format))
    if cell.font is not None and cell.font.bold:
        return text_len + max(BOLD_MIN_BONUS, text_len * (BOLD_WIDTH_FACTOR - 1))
    return text_len


def autofit_worksheet_columns(
    ws: Worksheet,
    get_measured_value: Callable[[Cell], object] = lambda cell: cell.value,
    min_width: float = MIN_COLUMN_WIDTH,
    max_width: float = MAX_COLUMN_WIDTH,
    column_max_width: Optional[Dict[int, float]] = None,
) -> None:
    """Set every column's width on `ws` to fit the widest VISIBLE
    content actually in it -- header text and data values -- mirroring
    Excel's own AutoFit Column Width, bounded to [`min_width`,
    `max_width`] by default.

    `get_measured_value` resolves what should actually be MEASURED for
    a given cell, defaulting to its raw `.value` -- callers whose
    cells may hold a live formula (whose `.value` in openpyxl is the
    FORMULA TEXT, e.g. `"=D4-E4"`, not what a person actually sees)
    should pass a callback that resolves the cell's already-computed
    value instead (e.g. via `SummaryWriter._formula_cache`, or a
    `data_only=True` sibling load of the same sheet).

    `column_max_width`, if given, maps a 1-based column index to a
    DIFFERENT max width than the general `max_width` -- used so the
    Comments column can be wider than other columns (see
    `MAX_COMMENTS_COLUMN_WIDTH`), giving wrapped text more room per
    line so it needs fewer lines, which is what keeps every row the
    same height regardless of comment length.
    """
    widest_by_col: Dict[int, float] = {}

    for row in ws.iter_rows():
        for cell in row:
            value = get_measured_value(cell)
            if value is None:
                continue
            text_len = measured_length(cell, value)
            if text_len > widest_by_col.get(cell.column, 0):
                widest_by_col[cell.column] = text_len

    overrides = column_max_width or {}
    for col, widest in widest_by_col.items():
        this_max = overrides.get(col, max_width)
        ws.column_dimensions[get_column_letter(col)].width = max(min(widest + 2, this_max), min_width)
