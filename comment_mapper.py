"""
comment_mapper.py
==================
Implements Rule 5: comments come from the "<year>_ClientComments" sheet.

That sheet has four columns:
    Group        -> the Sub-Group / DS-code the comment belongs to
                    (e.g. "DS10_Secured"). NOTE: this is a different
                    "Group" than the main sheet's customer-name Group --
                    confusingly, the comments sheet's "Group" column
                    actually stores the *section code*.
    Client List  -> the customer/group display name to match against our
                    aggregated Group (this is what Rule 5 calls "Group").
    Confidence   -> renewal confidence (not used in the Summary output).
    Comments     -> the free-text comment to copy into the Summary.

We match on (DS-code, normalised Client List) when possible, which
disambiguates cases where the same customer name appears in more than one
section (e.g. "HPE" exists under both DS10_Secured and DS70_Secured with
different comments). If an exact (code, name) match isn't found we fall
back to a name-only match and flag it, rather than silently leaving the
comment blank.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from typing import Dict, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet

from excel_reader import normalize_header, normalize_name, is_blank, extract_ds_code

logger = logging.getLogger("sfae.comment_mapper")


@dataclass
class CommentEntry:
    ds_code: Optional[int]
    raw_group_code: Optional[str]
    client_name: str
    confidence: Optional[str]
    comment: Optional[str]


class CommentMapper:
    def __init__(self, ws: Optional[Worksheet]):
        self._exact: Dict[Tuple[Optional[int], str], CommentEntry] = {}
        self._by_name: Dict[str, list] = {}
        self.available = ws is not None
        if ws is not None:
            self._load(ws)

    # ------------------------------------------------------------------
    def _find_columns(self, ws: Worksheet) -> Dict[str, int]:
        header_row = None
        for row in range(1, 6):
            values = [normalize_header(ws.cell(row, c).value) for c in range(1, ws.max_column + 1)]
            if "clientlist" in values or "comments" in values:
                header_row = row
                break
        if header_row is None:
            raise ValueError(
                f"Could not locate a header row with a 'Client List' or 'Comments' "
                f"column in sheet '{ws.title}'."
            )
        cols: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            norm = normalize_header(ws.cell(header_row, c).value)
            if norm == "group":
                cols["group"] = c
            elif norm == "clientlist":
                cols["client_list"] = c
            elif norm == "confidence":
                cols["confidence"] = c
            elif norm == "comments":
                cols["comments"] = c
        cols["_header_row"] = header_row
        return cols

    def _load(self, ws: Worksheet) -> None:
        cols = self._find_columns(ws)
        if "client_list" not in cols:
            logger.warning(
                "Comments sheet '%s' has no 'Client List' column; comment matching disabled.",
                ws.title,
            )
            self.available = False
            return

        start = cols["_header_row"] + 1
        for r in range(start, ws.max_row + 1):
            client_val = ws.cell(r, cols["client_list"]).value
            if is_blank(client_val):
                continue
            group_val = ws.cell(r, cols["group"]).value if "group" in cols else None
            comment_val = ws.cell(r, cols["comments"]).value if "comments" in cols else None
            confidence_val = ws.cell(r, cols["confidence"]).value if "confidence" in cols else None

            entry = CommentEntry(
                ds_code=extract_ds_code(group_val),
                raw_group_code=(str(group_val).strip() if not is_blank(group_val) else None),
                client_name=str(client_val).strip(),
                confidence=(str(confidence_val).strip() if not is_blank(confidence_val) else None),
                comment=(str(comment_val).strip() if not is_blank(comment_val) else None),
            )
            key_name = normalize_name(entry.client_name)
            self._exact[(entry.ds_code, key_name)] = entry
            self._by_name.setdefault(key_name, []).append(entry)

    # ------------------------------------------------------------------
    def lookup(self, group_name: str, ds_code: Optional[int]) -> Tuple[Optional[str], str]:
        """Return (comment_text_or_None, method) where method is one of
        'exact', 'name_only_fallback', or 'not_found'."""
        if not self.available:
            return None, "not_found"

        key_name = normalize_name(group_name)
        entry = self._exact.get((ds_code, key_name))
        if entry is not None:
            return entry.comment, "exact"

        candidates = self._by_name.get(key_name)
        if candidates:
            # Fall back to a name-only match (first candidate) but make
            # the caller aware this wasn't a fully disambiguated match.
            return candidates[0].comment, "name_only_fallback"

        return None, "not_found"
