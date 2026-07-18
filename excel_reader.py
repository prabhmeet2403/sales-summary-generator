"""
excel_reader.py
================
Everything related to *finding things* inside the Master workbook:
sheets, header rows, and columns -- all done dynamically by pattern /
header text, never by a fixed index or letter (per project Rule 7).

This module has no knowledge of business rules (grouping, quarters,
margin, comments matching). It only turns raw worksheet cells into
clean, typed Python objects that the rest of the application can work
with.
"""
from __future__ import annotations

import re
import logging
from collections import Counter
from dataclasses import dataclass, field, replace
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

import config

logger = logging.getLogger("sfae.excel_reader")


# ==========================================================================
# Exceptions
# ==========================================================================
class SheetNotFoundError(Exception):
    """Raised when a required sheet cannot be located dynamically."""


class ColumnNotFoundError(Exception):
    """Raised when a required column cannot be located dynamically."""


# ==========================================================================
# Small text helpers
# ==========================================================================
def normalize_header(value) -> str:
    """Lower-case a header value and strip everything that isn't a-z0-9.

    This makes "Total Revenue", "Total  Revenue", and "Total (Revenue)"
    all compare equal, without needing a long list of literal aliases for
    every cosmetic variation a spreadsheet author might type.
    """
    if value is None:
        return ""
    text = str(value).lower()
    return re.sub(r"[^a-z0-9]", "", text)


def normalize_name(value) -> str:
    """Normalise a customer/group display name for matching purposes."""
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def as_number(value) -> float:
    """Best-effort conversion of a cell value to float, treating blanks
    and non-numeric junk as 0 (Rule 6: numeric blanks = 0)."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("$", "").strip()
        if cleaned in ("", "-"):
            return 0.0
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    return 0.0


# ==========================================================================
# Sheet discovery
# ==========================================================================
def find_sheet_by_pattern(
    wb: Workbook,
    pattern: str,
    year: Optional[int] = None,
    exclude_keywords: Optional[List[str]] = None,
) -> Optional[str]:
    """Return the best-matching sheet name for a regex pattern.

    If `year` is given, prefer a sheet whose captured year group equals
    it. Sheets containing any of `exclude_keywords` (case-insensitive)
    are skipped, so that backup/duplicate sheets are never chosen ahead
    of the primary one.
    """
    exclude_keywords = [k.lower() for k in (exclude_keywords or [])]
    candidates: List[Tuple[str, Optional[int]]] = []
    for name in wb.sheetnames:
        if any(k in name.lower() for k in exclude_keywords):
            continue
        m = re.search(pattern, name, flags=re.IGNORECASE)
        if m:
            found_year = int(m.group(1)) if m.groups() else None
            candidates.append((name, found_year))

    if not candidates:
        return None

    if year is not None:
        for name, found_year in candidates:
            if found_year == year:
                return name
        return None

    # No specific year requested: prefer the one with the highest year,
    # falling back to the first match if none carry a year.
    with_year = [c for c in candidates if c[1] is not None]
    if with_year:
        return max(with_year, key=lambda c: c[1])[0]
    return candidates[0][0]


def discover_years_for_pattern(wb: Workbook, pattern: str) -> List[int]:
    """Return every year found in sheet names matching `pattern`, sorted
    ascending."""
    years = set()
    for name in wb.sheetnames:
        m = re.search(pattern, name, flags=re.IGNORECASE)
        if m and m.groups():
            years.add(int(m.group(1)))
    return sorted(years)


def _year_from_header_value(value: object) -> Optional[int]:
    """Pull a calendar year out of one header cell's value, whether
    it's a real Excel date/datetime (e.g. a "Jan" column literally
    storing 2026-01-01) or text written as a date (e.g. "Jan-26",
    "Jan 2026", "Q1 2026", "Q1'26"). Returns None if the value carries
    no recognizable year at all (plain labels like "Actual", "Total",
    "Name" etc. are expected to return None here)."""
    if isinstance(value, (datetime, date)):
        return value.year
    if isinstance(value, str):
        four_digit = re.search(r"\b(19|20)\d{2}\b", value)
        if four_digit:
            return int(four_digit.group(0))
        two_digit = re.search(r"[-'](\d{2})\b", value)
        if two_digit:
            return 2000 + int(two_digit.group(1))
    return None


def extract_business_year_from_header(ws: Worksheet) -> Optional[int]:
    """Determine the business year a sheet actually covers by reading
    its own header row's month/date columns -- e.g. a real Excel date
    like 2026-01-01 stored under a "Jan" column, or text like "Jan-26"
    -- never by parsing the sheet's own NAME. Reuses `find_header_rows`
    (the same header-row detection `build_column_map` itself relies on,
    which already specifically prefers whichever row actually holds
    real date-typed month columns over a coarser summary row above it)
    rather than assuming a fixed row number.

    Returns the single most common year found among the header's
    date-like cells (guarding against one stray/incorrect date cell
    outvoting the rest), or None if no header cell carries any
    recognizable year information at all.
    """
    field_row, type_row = find_header_rows(ws)
    max_col = min(ws.max_column, 80)
    year_counts: Counter = Counter()

    for row in {r for r in (field_row, type_row) if r}:
        for col in range(1, max_col + 1):
            year = _year_from_header_value(ws.cell(row=row, column=col).value)
            if year is not None:
                year_counts[year] += 1

    if not year_counts:
        return None
    return year_counts.most_common(1)[0][0]


# ==========================================================================
# Header detection
# ==========================================================================
@dataclass
class ColumnMap:
    """Resolved column positions (1-based) for one worksheet."""

    name: Optional[int] = None
    poc: Optional[int] = None
    service: Optional[int] = None
    group: Optional[int] = None
    sub_group: Optional[int] = None
    comments: Optional[int] = None
    renewal_confidence: Optional[int] = None
    total_revenue: Optional[int] = None
    total_margin: Optional[int] = None
    # month -> {"revenue": col, "salary": col, "margin": col}
    months: Dict[int, Dict[str, int]] = field(default_factory=dict)
    field_header_row: int = 0
    type_header_row: int = 0
    # Trailing "reference" columns embedded directly in the current
    # year's sheet that carry a prior year's figure for context, e.g. a
    # bare "2024" header, or a "2025_Total" header. Both are collapsed
    # into a single {year: column} map since they serve the same purpose
    # (a same-sheet snapshot of that year's Total) -- see
    # historical_lookup.py for why this turned out to be the primary
    # source of truth for the Summary's historical columns.
    year_reference_totals: Dict[int, int] = field(default_factory=dict)
    # A same-sheet "2025_Q4"-style reference: just that year's Q4
    # (Oct+Nov+Dec) Total, used purely as an independent drift detector
    # (see aggregator.attach_historical) -- if a fresh Q4-only
    # recomputation from that year's own sheet agrees with this
    # reference, the underlying sheet hasn't been revised since the
    # reference was captured, which tells us whether to trust a fresh
    # full-year recomputation over the (possibly mistyped) full-year
    # reference.
    year_reference_q4: Dict[int, int] = field(default_factory=dict)


def find_header_rows(ws: Worksheet, max_scan_rows: int = 20) -> Tuple[int, int]:
    """Locate the field-header row (the one containing literal "Name",
    "POC", "Group" etc.) and the type-header row directly above it (the
    one containing "Actual"/"Salary"/"Margin"/"Forecast" labels).

    Some workbooks repeat a coarse "Name"/"Group" style label on more
    than one header row (e.g. a short summary row above the fully
    detailed row that actually carries real Excel dates in its month
    columns). To avoid locking onto the wrong one, among every row that
    contains both "Name" and "Group" we prefer the row with the most
    actual date-typed cells (i.e. the one that really holds the 12
    month columns); if none contain any dates we fall back to the first
    matching row.

    Returns (field_header_row, type_header_row). type_header_row may be 0
    if there is no row above the field header row.
    """
    candidates: List[Tuple[int, int]] = []  # (row, count_of_date_cells)
    max_col = min(ws.max_column, 80)
    for row in range(1, max_scan_rows + 1):
        row_values = [ws.cell(row, c).value for c in range(1, max_col + 1)]
        normalized = [normalize_header(v) for v in row_values]
        if "name" in normalized and "group" in normalized:
            date_count = sum(1 for v in row_values if isinstance(v, (datetime, date)))
            candidates.append((row, date_count))

    if not candidates:
        raise ColumnNotFoundError(
            f"Could not locate a header row containing both 'Name' and 'Group' "
            f"in the first {max_scan_rows} rows of sheet '{ws.title}'."
        )

    best_row = max(candidates, key=lambda c: c[1])[0] if any(c[1] for c in candidates) else candidates[0][0]
    type_row = best_row - 1 if best_row > 1 else 0
    return best_row, type_row


def build_column_map(ws: Worksheet) -> ColumnMap:
    """Scan the header rows of `ws` and dynamically resolve every column
    this application cares about, by header text -- never by column
    letter or fixed index.
    """
    field_row, type_row = find_header_rows(ws)
    cmap = ColumnMap(field_header_row=field_row, type_header_row=type_row)

    max_col = ws.max_column
    for col in range(1, max_col + 1):
        field_val = ws.cell(field_row, col).value
        norm = normalize_header(field_val)

        # --- month columns (field header holds an actual date) ---------
        if isinstance(field_val, (datetime, date)):
            month = field_val.month
            type_val = ws.cell(type_row, col).value if type_row else None
            type_norm = normalize_header(type_val)
            role = None
            if any(k in type_norm for k in config.REVENUE_ROLE_KEYWORDS):
                role = "revenue"
            elif any(k in type_norm for k in config.SALARY_ROLE_KEYWORDS):
                role = "salary"
            elif any(k in type_norm for k in config.MARGIN_ROLE_KEYWORDS):
                role = "margin"
            if role:
                cmap.months.setdefault(month, {})[role] = col
            continue

        # --- plain text columns ----------------------------------------
        if norm == "name" and cmap.name is None:
            cmap.name = col
        elif norm == "poc" and cmap.poc is None:
            cmap.poc = col
        elif norm == "service" and cmap.service is None:
            cmap.service = col
        elif norm == "subgroup" and cmap.sub_group is None:
            cmap.sub_group = col
        elif norm == "group" and cmap.group is None:
            cmap.group = col
        elif norm == "comments" and cmap.comments is None:
            cmap.comments = col
        elif norm == "renewalconfidence" and cmap.renewal_confidence is None:
            cmap.renewal_confidence = col
        elif norm == "totalmargin" and cmap.total_margin is None:
            cmap.total_margin = col
        elif norm == "totalrevenue" and cmap.total_revenue is None:
            cmap.total_revenue = col
        elif norm == "total" and cmap.total_revenue is None:
            # Older-style sheets (e.g. a legacy year sheet with no
            # Actual/Salary/Margin split) may only have a bare "Total"
            # column, which represents total revenue.
            cmap.total_revenue = col
        elif re.fullmatch(r"\d{4}", norm):
            # A bare "2024"-style header: a same-sheet reference to that
            # year's Total.
            cmap.year_reference_totals[int(norm)] = col
        elif re.fullmatch(r"\d{4}total", norm):
            # A "2025_Total"-style header (any punctuation between the
            # year and "Total" is stripped by normalize_header): also a
            # same-sheet reference to that year's Total.
            year = int(norm[:4])
            cmap.year_reference_totals[year] = col
        elif re.fullmatch(r"\d{4}q4", norm):
            # A "2025_Q4"-style header: a same-sheet reference to that
            # year's Q4 only, used purely as a drift detector.
            year = int(norm[:4])
            cmap.year_reference_q4[year] = col

    if cmap.name is None:
        raise ColumnNotFoundError(
            f"Required column 'Name' not found in sheet '{ws.title}' "
            f"(header row {field_row})."
        )
    return cmap


# ==========================================================================
# Row-level data model
# ==========================================================================
@dataclass
class ProjectRow:
    """One project-level record from the main Sales-by-Customer sheet."""

    row_index: int
    name: str
    poc: Optional[str]
    group: str
    section_key: Optional[str]
    sub_group_raw: Optional[str]
    ds_code: Optional[int]
    monthly_revenue: Dict[int, float]   # month(1-12) -> revenue
    monthly_margin: Dict[int, float]    # month(1-12) -> margin
    sheet_total_revenue: Optional[float]
    sheet_total_margin: Optional[float]
    # year -> raw cell value from that year's embedded reference column
    # on THIS sheet (e.g. a "2024" or "2025_Total" column). The raw
    # value (not yet numeric-converted) is kept so callers can tell "the
    # cell was genuinely blank" apart from "the cell explicitly says 0 /
    # a dash", which matters for deciding whether to trust this
    # same-sheet snapshot or fall back to recomputing from that year's
    # own sheet.
    historical_refs_raw: Dict[int, object] = field(default_factory=dict)
    historical_q4_raw: Dict[int, object] = field(default_factory=dict)
    has_renewal_confidence: bool = False


def extract_ds_code(sub_group_value) -> Optional[int]:
    if is_blank(sub_group_value):
        return None
    m = re.search(config.DS_CODE_PATTERN, str(sub_group_value), flags=re.IGNORECASE)
    if not m:
        return None
    return int(m.group(1))


def read_project_rows(ws: Worksheet, cmap: ColumnMap) -> List[ProjectRow]:
    """Read every genuine data row from the sheet.

    A row is considered a genuine project row when it carries a non-blank
    Sub-Group value (section headers, subtotal rows, and blank spacer
    rows never do). If the sheet has no Sub-Group column at all, we fall
    back to "non-blank Group column" as the row filter.
    """
    rows: List[ProjectRow] = []
    start = cmap.field_header_row + 1
    current_section = None
    for r in range(start, ws.max_row + 1):

        first_cell = ws.cell(r, 1).value
        sub_group_val = ws.cell(r, cmap.sub_group).value if cmap.sub_group else None

        # Section-heading detection below only applies to rows that are
        # structurally headings -- every genuine heading/banner/subtotal
        # row on this sheet has a blank Sub-Group, while every genuine
        # data row has a populated one (verified across every
        # "Sales by Customer- <year>" sheet in every available master
        # workbook). Without this guard, a DATA row whose own Name
        # happens to contain one of these keywords -- e.g. "MetaSys
        # Staffing Reports Wendy", a genuine Investments-section row --
        # gets misread as if it were a new section heading, corrupting
        # `current_section` for itself and every row after it until the
        # next real heading. When there's no Sub-Group column at all
        # (`cmap.sub_group is None`), this is vacuously true for every
        # row, leaving that fallback path's behavior unchanged.
        is_heading_candidate = cmap.sub_group is None or is_blank(sub_group_val)

        if is_heading_candidate and isinstance(first_cell, str):
            text = first_cell.strip().lower()

            if "solutions and staff augmentation" in text:
                current_section = "projects_track1"

            elif "staffing" in text:
                current_section = "staffing_secured"

            elif "investment" in text:
                current_section = "investments"

            elif "track 1" in text and "projection" in text:
                current_section = "projects_track1_projection"

            elif "track 2" in text and "projection" in text:
                current_section = "projects_track2_projection"

        name_val = ws.cell(r, cmap.name).value
        group_val = ws.cell(r, cmap.group).value if cmap.group else None


        if cmap.sub_group:
            if is_blank(sub_group_val):
                continue
        else:
            if is_blank(group_val):
                continue

        if is_blank(name_val) or is_blank(group_val):
            # A row can carry a Sub-Group tag on a stray blank line in
            # some sheets; skip anything without both a Name and a Group.
            continue

        monthly_revenue: Dict[int, float] = {}
        monthly_margin: Dict[int, float] = {}
        for month, roles in cmap.months.items():
            if "revenue" in roles:
                monthly_revenue[month] = as_number(ws.cell(r, roles["revenue"]).value)
            if "margin" in roles:
                monthly_margin[month] = as_number(ws.cell(r, roles["margin"]).value)

        sheet_total_revenue = (
            as_number(ws.cell(r, cmap.total_revenue).value) if cmap.total_revenue else None
        )
        sheet_total_margin = (
            as_number(ws.cell(r, cmap.total_margin).value) if cmap.total_margin else None
        )
        historical_refs_raw = {
            year: ws.cell(r, col).value for year, col in cmap.year_reference_totals.items()
        }
        historical_q4_raw = {
            year: ws.cell(r, col).value for year, col in cmap.year_reference_q4.items()
        }
        renewal_val = ws.cell(r, cmap.renewal_confidence).value if cmap.renewal_confidence else None

        rows.append(
            ProjectRow(
                row_index=r,
                name=str(name_val).strip(),
                poc=(str(ws.cell(r, cmap.poc).value).strip() if cmap.poc and not is_blank(ws.cell(r, cmap.poc).value) else None),
                group=str(group_val).strip(),
                section_key=current_section,
                sub_group_raw=(str(sub_group_val).strip() if not is_blank(sub_group_val) else None),
                ds_code=extract_ds_code(sub_group_val),
                monthly_revenue=monthly_revenue,
                monthly_margin=monthly_margin,
                sheet_total_revenue=sheet_total_revenue,
                sheet_total_margin=sheet_total_margin,
                historical_refs_raw=historical_refs_raw,
                historical_q4_raw=historical_q4_raw,
                has_renewal_confidence=not is_blank(renewal_val),
            )
        )
    return rows


# ==========================================================================
# Top level convenience wrapper
# ==========================================================================
class MasterWorkbook:
    """Thin convenience wrapper around an opened Master workbook."""

    def __init__(self, path: str):
        self.path = path
        logger.info("Loading workbook: %s", path)
        self.wb = load_workbook(path, data_only=True)

    def main_sheet_name(self, year: Optional[int] = None) -> str:
        name = find_sheet_by_pattern(self.wb, config.MAIN_SHEET_PATTERN, year=year)
        if not name:
            available = ", ".join(self.wb.sheetnames)
            raise SheetNotFoundError(
                f"Could not find the main 'Sales by Customer' sheet"
                f"{f' for year {year}' if year else ''}. "
                f"Available sheets: {available}"
            )
        return name

    def comments_sheet_name(self, year: Optional[int] = None) -> Optional[str]:
        return find_sheet_by_pattern(
            self.wb,
            config.COMMENTS_SHEET_PATTERN,
            year=year,
            exclude_keywords=config.COMMENTS_SHEET_EXCLUDE_KEYWORDS,
        )

    def available_years(self) -> List[int]:
        return discover_years_for_pattern(self.wb, config.MAIN_SHEET_PATTERN)

    def detect_business_year_from_content(self) -> int:
        """Determine the DEFAULT business year -- used for the output
        filename and everywhere else `target_year` is needed -- from
        the main "Sales by Customer" sheet's own header content (its
        month/date columns), never by parsing a year out of any sheet's
        NAME.

        Locating WHICH sheet is the main data sheet still uses the
        existing name-pattern matching (`find_sheet_by_pattern`,
        unchanged, and still how a user's explicit `--year`/"Detected
        Year" override picks a specific prior-year sheet -- that
        override is a separate, pre-existing feature this doesn't
        touch) -- but the actual YEAR VALUE that sheet represents is
        then read from ITS OWN header, not from its name, satisfying
        "do not rely on the sheet name" for the year itself.

        Raises SheetNotFoundError with a clear message (rather than
        silently falling back to name-parsing or any other source) if
        no main sheet can be found at all, or if that sheet's header
        carries no recognizable year information.
        """
        main_name = find_sheet_by_pattern(self.wb, config.MAIN_SHEET_PATTERN, year=None)
        if not main_name:
            raise SheetNotFoundError(
                "No sheet matching 'Sales by Customer' was found in the workbook, "
                "so the business year could not be determined from its content. "
                f"Available sheets: {', '.join(self.wb.sheetnames)}"
            )
        try:
            year = extract_business_year_from_header(self.wb[main_name])
        except ColumnNotFoundError:
            # The main sheet doesn't even have the expected header
            # structure to scan (e.g. no "Name"/"Group" header row at
            # all) -- same clear failure as "no year found", just a
            # different reason for it.
            year = None
        if year is None:
            raise SheetNotFoundError(
                f"Could not determine the business year from sheet '{main_name}'s "
                "own header content. The business year is read from that sheet's "
                "month/date header columns (e.g. real dates or labels like "
                "'Jan-26'), not from the sheet's name -- please confirm the "
                "header row actually shows recognizable year information."
            )
        return year

    def sheet(self, name: str) -> Worksheet:
        if name not in self.wb.sheetnames:
            raise SheetNotFoundError(
                f"Sheet '{name}' not found. Available sheets: {', '.join(self.wb.sheetnames)}"
            )
        return self.wb[name]


# A section's row_range set to this exact sentinel means its own
# title text could not be found anywhere in this specific source
# workbook at all -- i.e. this section doesn't apply to this
# workbook's layout (as opposed to applying but genuinely having zero
# rows this year, which is a normal, different situation still shown
# with its usual heading/empty-subtotal banner). Callers building the
# final section list for output (main.py/gui/runner.py) should exclude
# any section carrying this sentinel entirely, rather than showing an
# empty heading/subtotal pair for a section that isn't part of this
# workbook at all.
NO_MATCHING_ROWS: Tuple[int, int] = (-2, -1)


def disambiguate_shared_ds_code_sections(
    ws: Worksheet, sections: List["config.OutputSection"], name_col: int = 1,
) -> List["config.OutputSection"]:
    """Return a COPY of `sections` where any section whose `ds_codes`
    overlap with another section's `ds_codes` gets its `row_range`
    (see `config.OutputSection`) filled in dynamically, from `ws`'s own
    row positions -- so two sections that legitimately reuse the same
    DS-code (e.g. "Investments" and "projects_track1" both use
    DS10_Secured in the Master workbook) don't each pull in the
    other's rows too. Sections whose `ds_codes` are unique are
    returned completely unchanged.

    For each section needing disambiguation, its own `title` text is
    searched for verbatim in the sheet's own Name column (`name_col`,
    resolved dynamically from that sheet's header via `build_column_map`
    -- never hardcoded to column A, since a differently-ordered source
    sheet may have Name anywhere) to find where its block starts on
    THIS specific workbook; its row_range runs from just after that row
    to just before whichever OTHER disambiguated section's own title
    row comes next (in `sections`' own order), or to the end of the
    sheet if none does. Never a hardcoded pair of row numbers -- purely
    derived from what's actually in `ws`.

    A section whose own title text can't be found in `ws` at all (a
    different source workbook's layout, missing that section entirely)
    gets a row_range that can never match any real row, rather than
    being left unrestricted -- so it safely produces zero rows instead
    of incorrectly claiming a DS-code sibling's rows once that
    sibling's own disambiguating range is applied.
    """
    code_counts: Dict[int, int] = {}
    for section in sections:
        for code in section.ds_codes:
            code_counts[code] = code_counts.get(code, 0) + 1
    needs_disambiguation = [s for s in sections if any(code_counts[c] > 1 for c in s.ds_codes)]
    if not needs_disambiguation:
        return sections

    title_rows: Dict[str, Optional[int]] = {
        section.key: _find_row_by_exact_label(ws, section.title, name_col) for section in needs_disambiguation
    }

    result: List["config.OutputSection"] = []
    for section in sections:
        if section.key not in title_rows:
            result.append(section)
            continue
        start_row = title_rows[section.key]
        if start_row is None:
            result.append(replace(section, row_range=NO_MATCHING_ROWS))
            continue
        end_row = ws.max_row
        for other in needs_disambiguation:
            if other.key == section.key:
                continue
            other_row = title_rows.get(other.key)
            if other_row is not None and other_row > start_row:
                end_row = min(end_row, other_row - 1)
        result.append(replace(section, row_range=(start_row + 1, end_row)))
    return result


def _find_row_by_exact_label(ws: Worksheet, label: str, name_col: int = 1) -> Optional[int]:
    """Row index (1-based) of the first cell in column `name_col`
    whose text exactly matches `label` (after stripping surrounding
    whitespace), or None if not found anywhere on `ws`."""
    target = label.strip()
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row=row, column=name_col).value or "").strip() == target:
            return row
    return None
