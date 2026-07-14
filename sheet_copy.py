"""
sheet_copy.py
================
Appends a visual copy of one worksheet from the uploaded Master
workbook onto the generated Summary workbook, as a third sheet -- used
for the "Sales by Customer- <year>" raw-data sheet requested alongside
the two generated (aggregated) sheets `summary_writer.py` already
produces.

This is a distinct operation from everything in `summary_writer.py`:
that module turns already-aggregated `GroupSummary` objects into new,
formatted rows it designs itself. This module does the opposite -- it
copies a worksheet that already exists, cell for cell, style for
style, exactly as the source workbook has it, with one column (the
Comments column, located dynamically by its header text -- never a
hardcoded letter) physically removed and every column after it shifted
left by one, mirroring exactly what Excel's own "Delete Column" does.
Kept separate so neither module's job is duplicated inside the other.

FORMULAS: same-sheet formulas stay live; cross-sheet formulas become
values. A formula like `=D4-E4` or `=SUMPRODUCT(D4:AK4)` only ever
needs cells on THIS sheet, which this module also copies in full, so
it keeps recalculating correctly and is kept as a live formula (with
its own cell references shifted for the column removal, exactly as
Excel itself would on a real "Delete Column"). A formula like
`=VLOOKUP($A13,'Salary Projections 2026'!$A:$D,3,FALSE)` depends on a
DIFFERENT sheet that this module does not also bring over (only the
one sheet named is copied) -- Excel cannot resolve that sheet inside
this workbook, treats it as needing an external source, and shows both
the "this workbook contains links to external sources" warning and a
#REF!/#N/A in the cell. Those are converted to their already-computed
value instead (read from a `data_only=True` load of the source -- i.e.
exactly what Excel itself last displayed there, never recomputed by
this code), which carries no formula for Excel to fail to resolve.
"""

from __future__ import annotations

import colorsys
import datetime
import re
import xml.etree.ElementTree as ET
from copy import copy, deepcopy
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles.colors import Color
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

# Excel's theme-color index order for a `<color theme="N">` reference --
# NOT the raw document order the <a:clrScheme> XML itself lists colors
# in (dk1, lt1, dk2, lt2, accent1-6, hlink, folHlink). The first two
# slots are swapped relative to that document order; this exact swap is
# a well-documented, widely-relied-on quirk of the OOXML spreadsheet
# theme-color convention.
_THEME_SLOT_ORDER = (
    "lt1", "dk1", "lt2", "dk2",
    "accent1", "accent2", "accent3", "accent4", "accent5", "accent6",
    "hlink", "folHlink",
)
_DRAWINGML_NS = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}


def _parse_theme_palette(theme_xml: Optional[bytes]) -> Optional[List[str]]:
    """Return the workbook's 12-slot theme color palette, as 6-digit
    hex RGB strings in Excel's own theme-index order (see
    `_THEME_SLOT_ORDER`) -- or None if `theme_xml` is missing or
    doesn't parse, in which case theme colors are left unresolved
    (copied through as-is, same as before this fix existed).
    """
    if not theme_xml:
        return None
    try:
        root = ET.fromstring(theme_xml)
    except ET.ParseError:
        return None
    scheme = root.find(".//a:clrScheme", _DRAWINGML_NS)
    if scheme is None:
        return None

    def _slot_hex(tag: str) -> Optional[str]:
        el = scheme.find(f"a:{tag}", _DRAWINGML_NS)
        if el is None:
            return None
        srgb = el.find("a:srgbClr", _DRAWINGML_NS)
        if srgb is not None:
            return srgb.get("val")
        sys_clr = el.find("a:sysClr", _DRAWINGML_NS)
        if sys_clr is not None:
            return sys_clr.get("lastClr")
        return None

    palette = [_slot_hex(tag) for tag in _THEME_SLOT_ORDER]
    return palette if all(palette) else None


def _apply_tint(hex_rgb: str, tint: float) -> str:
    """Excel's own tint/shade algorithm: convert sRGB to HSL, adjust
    luminance by `tint` (positive lightens toward white, negative
    darkens toward black), convert back. This is the same formula
    Excel itself uses to render a themed color's tinted variants (the
    lighter/darker swatches in the color picker), applied here so a
    resolved theme color LOOKS like what the source intended, not just
    the raw, untinted theme color.
    """
    r, g, b = (int(hex_rgb[i:i + 2], 16) / 255 for i in (0, 2, 4))
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    if tint < 0:
        l = l * (1.0 + tint)
    else:
        l = l * (1.0 - tint) + tint
    l = min(1.0, max(0.0, l))
    r2, g2, b2 = colorsys.hls_to_rgb(h, l, s)
    return f"{round(r2 * 255):02X}{round(g2 * 255):02X}{round(b2 * 255):02X}"


def _resolve_theme_color(color: Optional[Color], theme_palette: Optional[List[str]]) -> Optional[Color]:
    """If `color` is a theme-indexed reference (`type == "theme"`),
    return a NEW `Color` with an explicit, resolved `rgb` value instead
    -- otherwise return `color` unchanged (already explicit rgb,
    legacy indexed, or auto).

    This is the actual fix for a real bug: a theme color's rendered
    RGB depends on the WORKBOOK'S OWN theme definition
    (`xl/theme/theme1.xml`), which differs between the uploaded Master
    workbook and a freshly-created `openpyxl.Workbook()` (the latter
    ships its own default "Office" theme). The same `theme=9` reference
    is `accent6`, which is a GREEN (`4EA72E`) in the Master workbook
    observed here but an ORANGE (`F79646`) in openpyxl's default theme
    -- identical theme index, genuinely different colors, which is
    exactly the wrong-color symptom this resolves. Converting to an
    explicit RGB value makes the color workbook-independent: it will
    render identically no matter what theme is active for whichever
    workbook this cell ends up in.

    Sheets 1/2 are entirely unaffected by this: they are built
    separately (`summary_writer.py`) and never touched here; resolving
    colors on THIS module's copied cells changes nothing about the
    workbook's own theme.xml or any other sheet's styling.
    """
    if color is None or color.type != "theme" or theme_palette is None:
        return color
    if color.theme < 0 or color.theme >= len(theme_palette):
        return color
    base_hex = theme_palette[color.theme]
    resolved_hex = _apply_tint(base_hex, color.tint or 0.0) if color.tint else base_hex
    return Color(rgb=f"FF{resolved_hex}")


def copy_source_sheet_as_new_worksheet(
    wb: Workbook,
    source_path: str,
    sheet_name: str,
    comments_col: Optional[int],
    formula_cache: List[Tuple[str, str, float]],
    output_sheet_name: Optional[str] = None,
) -> None:
    """Append `sheet_name` from `source_path` onto `wb` as a new sheet
    (titled `output_sheet_name`, or `sheet_name` itself if not given),
    copying everything needed to make it LOOK identical -- computed
    cell values or, where safe, live formulas (see module docstring),
    fonts/fills/borders/number formats/alignment, column widths, row
    heights, merged cells, freeze panes/split panes/scroll position,
    autofilter, conditional formatting, print/page setup, and hidden
    rows/columns -- with one exception: if `comments_col` is given
    (1-based column index, e.g. `ColumnMap.comments` -- resolved
    dynamically by header text elsewhere, never hardcoded), that
    ENTIRE column is removed (not just blanked), and every column
    after it shifts left by one, exactly like using Excel's own
    "Delete Column" on the source sheet.

    Args:
        wb: The in-progress output workbook (already has its other
            sheet(s) from `SummaryWriter.build()`) to append to.
        source_path: Path to the uploaded Master workbook.
        sheet_name: Exact sheet name to copy FROM in the source
            workbook (how the sheet is located there -- unaffected by
            whatever the new sheet ends up being called).
        comments_col: 1-based column index of the Comments column on
            the SOURCE sheet, or None if it has none (nothing is
            removed in that case).
        formula_cache: The list `SummaryWriter.patch_cached_formula_values`
            reads after `wb.save(...)` -- pass `writer._formula_cache`
            so this sheet's preserved same-sheet formulas get a cached
            value through that exact same, already-tested pass, with
            no separate injection step of its own.
        output_sheet_name: What to title the new sheet in `wb` --
            defaults to `sheet_name` (mirroring the source's own name)
            if not given. Kept separate so the OUTPUT workbook's sheet
            name can differ from the source's own (e.g. renaming this
            sheet to "<year> SOW Performance" while still reading from
            the source's "Sales by Customer- <year>" sheet) without
            affecting which sheet gets read.
    """
    output_sheet_name = output_sheet_name or sheet_name
    source_wb_values = load_workbook(source_path, data_only=True)
    source_wb_styles = load_workbook(source_path, data_only=False)
    src_values = source_wb_values[sheet_name]
    src_styles = source_wb_styles[sheet_name]
    theme_palette = _parse_theme_palette(source_wb_styles.loaded_theme)

    new_ws = wb.create_sheet(title=output_sheet_name)
    _copy_cells(src_values, src_styles, new_ws, comments_col, theme_palette, output_sheet_name, formula_cache)
    _copy_dimensions(src_styles, new_ws, comments_col)
    _autofit_column_widths(src_values, new_ws, comments_col)
    _copy_merged_cells(src_styles, new_ws, comments_col)
    _copy_sheet_view(src_styles, new_ws, comments_col)
    _copy_page_setup(src_styles, new_ws, comments_col)
    _copy_conditional_formatting(src_styles, new_ws, comments_col, theme_palette)


def _remap_column(col_idx: int, removed_col: Optional[int]) -> Optional[int]:
    """1-based column index after removing `removed_col` -- unchanged
    if before it, shifted left by one if after it, or None if it IS
    the removed column. Returns `col_idx` unchanged if `removed_col`
    is None (nothing being removed)."""
    if removed_col is None:
        return col_idx
    if col_idx == removed_col:
        return None
    if col_idx < removed_col:
        return col_idx
    return col_idx - 1


def _remap_range_string(range_str: Optional[str], removed_col: Optional[int]) -> Optional[str]:
    """Shift a range string to account for `removed_col` being deleted.
    Handles a single cell (`"D70"`), a single range (`"A2:AO56"`), a
    full-column range (`"AO1:AO1048576"`), and a MULTI-range string --
    conditional formatting's `sqref` in particular can be several
    space-separated ranges/cells covering disjoint areas (e.g.
    `"E16 E18 F51:F67 ..."`), each remapped independently here. Returns
    None if every piece was entirely inside the removed column (nothing
    left of the whole thing to keep).
    """
    if removed_col is None or not range_str:
        return range_str

    if " " in range_str:
        remapped_pieces = [
            piece for piece in (
                _remap_single_range(token, removed_col) for token in range_str.split()
            ) if piece is not None
        ]
        return " ".join(remapped_pieces) if remapped_pieces else None

    return _remap_single_range(range_str, removed_col)


def _remap_single_range(range_str: str, removed_col: int) -> Optional[str]:
    """Remap exactly one cell or range token (no spaces) -- see
    `_remap_range_string`, which this implements the single-piece case
    for."""
    if ":" not in range_str:
        col_str, row = coordinate_from_string(range_str)
        new_col = _remap_column(column_index_from_string(col_str), removed_col)
        if new_col is None:
            return None
        return f"{get_column_letter(new_col)}{row}"

    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    if min_col == max_col == removed_col:
        return None
    new_min = min_col if min_col < removed_col else max(min_col - 1, 1)
    new_max = max_col if max_col < removed_col else max(max_col - 1, 1)
    return f"{get_column_letter(new_min)}{min_row}:{get_column_letter(new_max)}{max_row}"


# Matches an A1-style cell reference with optional absolute-reference
# `$` markers (e.g. `A1`, `$A$1`, `$A1`, `A$1`). Used both to detect a
# cross-sheet formula (a `'Sheet Name'!` or `SheetName!` prefix right
# before a match) and to rewrite same-sheet formulas' own references
# for the column removal.
_CELL_REF_RE = re.compile(r"(\$?)([A-Za-z]{1,3})(\$?)(\d+)")
_SHEET_QUALIFIED_RE = re.compile(r"(?:'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!")
_STRING_LITERAL_RE = re.compile(r'"[^"]*"')


def _formula_references_other_sheet(formula: str) -> bool:
    """True if `formula` contains a sheet-qualified reference
    (`'Some Sheet'!A1` or `SheetName!A1`) anywhere outside a quoted
    string literal -- i.e. it depends on a DIFFERENT worksheet, which
    this module does not also copy. See the module docstring for why
    such formulas are converted to their value instead of kept live.
    """
    string_spans = [m.span() for m in _STRING_LITERAL_RE.finditer(formula)]
    for m in _SHEET_QUALIFIED_RE.finditer(formula):
        if not any(start <= m.start() < end for start, end in string_spans):
            return True
    return False


def _shift_formula_columns(formula: str, comments_col: Optional[int]) -> Optional[str]:
    """Rewrite every cell/range reference's column letter in a
    SAME-SHEET formula to account for `comments_col` being removed --
    mirroring exactly what Excel itself does when a column is deleted:
    references before it are untouched, references after it shift left
    by one. Skips any match inside a quoted string literal (so e.g. a
    label like `="Q1 Total: "&SUM(...)` doesn't have "Q1" mistaken for
    a cell reference). Returns None if the formula references the
    removed column itself -- which would become a dangling #REF! once
    that column no longer exists -- so the caller can fall back to the
    cell's value instead of keeping a broken formula.
    """
    if comments_col is None:
        return formula

    string_spans = [m.span() for m in _STRING_LITERAL_RE.finditer(formula)]
    broke_reference = False

    def _replace(m: re.Match) -> str:
        nonlocal broke_reference
        if any(start <= m.start() < end for start, end in string_spans):
            return m.group(0)
        dollar_col, col_letters, dollar_row, row_num = m.groups()
        try:
            col_idx = column_index_from_string(col_letters.upper())
        except ValueError:
            return m.group(0)
        if col_idx == comments_col:
            broke_reference = True
            return m.group(0)
        new_idx = col_idx if col_idx < comments_col else col_idx - 1
        return f"{dollar_col}{get_column_letter(new_idx)}{dollar_row}{row_num}"

    rewritten = _CELL_REF_RE.sub(_replace, formula)
    return None if broke_reference else rewritten


def _copy_cells(
    src_values: Worksheet,
    src_styles: Worksheet,
    dest: Worksheet,
    comments_col: Optional[int],
    theme_palette: Optional[List[str]],
    sheet_name: str,
    formula_cache: List[Tuple[str, str, float]],
) -> None:
    for row in src_styles.iter_rows():
        for cell in row:
            new_col = _remap_column(cell.column, comments_col)
            if new_col is None:
                continue  # this cell is in the removed Comments column

            new_cell = dest.cell(row=cell.row, column=new_col)
            _set_cell_value_or_formula(
                cell, new_cell, src_values, comments_col, sheet_name, formula_cache,
            )

            if cell.has_style:
                new_font = copy(cell.font)
                new_font.color = _resolve_theme_color(cell.font.color, theme_palette)
                new_cell.font = new_font

                new_cell.border = copy(cell.border)

                new_fill = copy(cell.fill)
                new_fill.fgColor = _resolve_theme_color(cell.fill.fgColor, theme_palette)
                new_fill.bgColor = _resolve_theme_color(cell.fill.bgColor, theme_palette)
                new_cell.fill = new_fill

                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)


def _set_cell_value_or_formula(
    cell,
    new_cell,
    src_values: Worksheet,
    comments_col: Optional[int],
    sheet_name: str,
    formula_cache: List[Tuple[str, str, float]],
) -> None:
    """Decide, for one cell, whether to keep a live (column-shifted)
    formula or write its computed value -- see the module docstring's
    "FORMULAS" section for the full reasoning.
    """
    raw_value = cell.value
    is_array_formula = isinstance(raw_value, ArrayFormula)
    formula_text = raw_value.text if is_array_formula else raw_value
    is_formula = isinstance(formula_text, str) and formula_text.startswith("=")

    cached_value = src_values.cell(row=cell.row, column=cell.column).value

    if not is_formula or _formula_references_other_sheet(formula_text):
        # Not a formula at all, OR depends on a sheet this module
        # doesn't also copy: write the already-computed value.
        new_cell.value = cached_value
        return

    shifted_text = _shift_formula_columns(formula_text, comments_col)
    if shifted_text is None:
        # References the Comments column itself -- would be a
        # dangling #REF! once that column is gone. Fall back to value.
        new_cell.value = cached_value
        return

    if is_array_formula:
        shifted_ref = _remap_range_string(raw_value.ref, comments_col) or new_cell.coordinate
        new_cell.value = ArrayFormula(ref=shifted_ref, text=shifted_text)
    else:
        new_cell.value = shifted_text

    # Same as every other live formula this app writes: queue the
    # already-computed result for `SummaryWriter.patch_cached_formula_values`
    # (see summary_writer.py) to inject after `wb.save(...)`, so the
    # formula displays correctly immediately rather than only after
    # Excel recalculates it.
    if isinstance(cached_value, (int, float)):
        formula_cache.append((sheet_name, new_cell.coordinate, float(cached_value)))


def _copy_dimensions(src: Worksheet, dest: Worksheet, comments_col: Optional[int]) -> None:
    for key, dim in src.column_dimensions.items():
        try:
            old_idx = column_index_from_string(key)
        except ValueError:
            continue  # not a plain column letter (e.g. a column-group marker) -- skip
        new_idx = _remap_column(old_idx, comments_col)
        if new_idx is None:
            continue  # the removed Comments column's own dimension
        new_dim = dest.column_dimensions[get_column_letter(new_idx)]
        # width is deliberately NOT copied here -- see
        # `_autofit_column_widths`, called separately after this, which
        # sets it based on actual cell content instead (an explicit,
        # separate request for this sheet specifically).
        new_dim.hidden = dim.hidden
        new_dim.outline_level = dim.outline_level
        new_dim.bestFit = dim.bestFit

    for key, dim in src.row_dimensions.items():
        new_dim = dest.row_dimensions[key]
        new_dim.height = dim.height
        new_dim.hidden = dim.hidden
        new_dim.outline_level = dim.outline_level

    dest.sheet_format.defaultColWidth = src.sheet_format.defaultColWidth
    dest.sheet_format.defaultRowHeight = src.sheet_format.defaultRowHeight


def _display_text(value, number_format: Optional[str]) -> str:
    """Approximate what Excel would actually SHOW for one cell, given
    its value and number format -- used only to measure text length
    for `_autofit_column_widths`, never written anywhere. Covers the
    number formats actually observed on this sheet (currency, accounting,
    thousands-separated plain numbers, dates, text, General) with a
    plain `str(value)` fallback for anything else -- a full general-
    purpose number-format renderer is out of scope for a width estimate,
    and openpyxl has no built-in AutoFit to defer to.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%b-%y")
    if isinstance(value, (int, float)):
        nf = number_format or "General"
        if nf in ("General", "@"):
            if isinstance(value, float) and value == int(value):
                return str(int(value))
            return str(value)
        if "$" in nf:
            if value < 0:
                return f"(${abs(value):,.0f})" if "(" in nf else f"-${abs(value):,.0f}"
            return f"${value:,.0f}"
        if "%" in nf:
            return f"{value * 100:.0f}%"
        if "#,##0" in nf:
            return f"{value:,.0f}"
        return str(value)
    return str(value)


def _autofit_column_widths(src_values: Worksheet, dest: Worksheet, comments_col: Optional[int]) -> None:
    """Set every column's width on `dest` based on the widest visible
    content actually in that column -- header text, data values, and
    (for formula cells) the DISPLAYED/computed result rather than the
    formula text -- mirroring Excel's own AutoFit Column Width.

    Reads from `src_values` (the `data_only=True` sibling load of the
    same source sheet already used elsewhere in this module) rather
    than `dest` itself, specifically so formula cells are measured by
    their computed result: at the point this runs, a preserved
    same-sheet formula cell on `dest` holds the FORMULA TEXT (e.g.
    `"=D4-E4"`), which is not what a person looking at the sheet would
    actually see -- `src_values` already has the real, already-computed
    number for that exact cell.

    This replaces `_copy_dimensions`'s usual "copy the source's own
    explicit width" behavior for this sheet specifically (an explicit,
    separate request) -- row heights and every other formatting
    attribute are untouched.
    """
    max_col = src_values.max_column
    max_row = src_values.max_row
    widest_by_new_col: Dict[int, int] = {}

    for col in range(1, max_col + 1):
        new_col = _remap_column(col, comments_col)
        if new_col is None:
            continue
        widest = widest_by_new_col.get(new_col, 0)
        for row in range(1, max_row + 1):
            cell = src_values.cell(row=row, column=col)
            if cell.value is None:
                continue
            text_len = len(_display_text(cell.value, cell.number_format))
            if text_len > widest:
                widest = text_len
        widest_by_new_col[new_col] = widest

    for new_col, widest in widest_by_new_col.items():
        # Excel's own column-width unit is approximately "number of
        # characters of the default font that fit", with a small
        # padding allowance for cell margins -- the same widely-used
        # approximation the openpyxl community relies on, since
        # openpyxl has no built-in AutoFit to call instead. A modest
        # floor keeps genuinely empty/near-empty columns from
        # collapsing to an unreadably thin sliver.
        dest.column_dimensions[get_column_letter(new_col)].width = max(widest + 2, 6)


def _copy_merged_cells(src: Worksheet, dest: Worksheet, comments_col: Optional[int]) -> None:
    for merged_range in src.merged_cells.ranges:
        remapped = _remap_range_string(str(merged_range), comments_col)
        if remapped is not None and ":" in remapped:  # a single-cell result isn't a merge
            dest.merge_cells(remapped)


def _copy_sheet_view(src: Worksheet, dest: Worksheet, comments_col: Optional[int]) -> None:
    """Copy freeze panes, split panes, scroll position, zoom, and
    selection -- everything governing how the sheet scrolls and what's
    visible on open -- by deep-copying the source's own `SheetView`
    object wholesale and only then fixing up the handful of fields
    that are actual cell/column references, rather than going through
    `Worksheet.freeze_panes`'s convenience setter.

    That setter is LOSSY for a pane that's both frozen and
    independently scrolled: its only input is `pane.topLeftCell`, which
    it (incorrectly, for this purpose) assumes IS the freeze split
    point. In the real Master workbook this sheet is copied from, the
    freeze split is at column D / row 3 (`xSplit=3`, `ySplit=2`), but
    the sheet had additionally been scrolled down before saving, so
    `topLeftCell` is `"D70"` -- merely where the scrollable pane's
    view happens to start, not the split boundary. Round-tripping
    `freeze_panes = "D70"` through the setter reinterprets row 70 AS
    the split point, freezing the top 69 rows instead of 2 -- an
    enormous, wrong frozen region, which is exactly what was breaking
    scrolling on the copied sheet.
    """
    dest.views.sheetView[0] = deepcopy(src.sheet_view)

    pane = dest.sheet_view.pane
    if pane is not None:
        if pane.topLeftCell:
            pane.topLeftCell = _remap_range_string(pane.topLeftCell, comments_col)
        if comments_col is not None and pane.xSplit:
            # xSplit is a COUNT of frozen columns from the left, not a
            # column reference -- if the removed column was itself
            # among the frozen columns, there is now one fewer.
            if comments_col <= pane.xSplit:
                pane.xSplit = pane.xSplit - 1

    for selection in dest.sheet_view.selection:
        if selection.activeCell:
            selection.activeCell = _remap_range_string(selection.activeCell, comments_col)
        if selection.sqref:
            selection.sqref = _remap_range_string(selection.sqref, comments_col)


def _copy_page_setup(src: Worksheet, dest: Worksheet, comments_col: Optional[int]) -> None:
    if src.auto_filter and src.auto_filter.ref:
        remapped_ref = _remap_range_string(src.auto_filter.ref, comments_col)
        if remapped_ref:
            dest.auto_filter.ref = remapped_ref

    dest.page_setup = copy(src.page_setup)
    dest.page_margins = copy(src.page_margins)
    dest.print_options = copy(src.print_options)
    if src.print_area:
        dest.print_area = _remap_range_string(src.print_area, comments_col)
    if src.print_title_cols:
        dest.print_title_cols = src.print_title_cols
    if src.print_title_rows:
        dest.print_title_rows = src.print_title_rows
    dest.oddHeader = copy(src.oddHeader)
    dest.oddFooter = copy(src.oddFooter)
    dest.evenHeader = copy(src.evenHeader)
    dest.evenFooter = copy(src.evenFooter)
    dest.firstHeader = copy(src.firstHeader)
    dest.firstFooter = copy(src.firstFooter)
    dest.sheet_properties.pageSetUpPr = copy(src.sheet_properties.pageSetUpPr)


def _copy_conditional_formatting(
    src: Worksheet, dest: Worksheet, comments_col: Optional[int], theme_palette: Optional[List[str]],
) -> None:
    for cf_range in src.conditional_formatting:
        remapped_sqref = _remap_range_string(str(cf_range.sqref), comments_col)
        if not remapped_sqref:
            continue  # this rule applied only to the now-removed column
        for rule in cf_range.rules:
            new_rule = copy(rule)
            dxf = new_rule.dxf
            if dxf is not None:
                if dxf.font is not None and dxf.font.color is not None:
                    dxf.font = copy(dxf.font)
                    dxf.font.color = _resolve_theme_color(dxf.font.color, theme_palette)
                if dxf.fill is not None:
                    dxf.fill = copy(dxf.fill)
                    if dxf.fill.fgColor is not None:
                        dxf.fill.fgColor = _resolve_theme_color(dxf.fill.fgColor, theme_palette)
                    if dxf.fill.bgColor is not None:
                        dxf.fill.bgColor = _resolve_theme_color(dxf.fill.bgColor, theme_palette)
            dest.conditional_formatting.add(remapped_sqref, new_rule)
