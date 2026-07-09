r"""
summary_writer.py
==================
Builds the final Sales & Forecast Summary workbook from the aggregated
GroupSummary objects produced by aggregator.py.

Column layout (fully dynamic based on config.NUM_PRIOR_YEARS_SHOWN /
config.YEARS_WITH_MARGIN_SHOWN):

    Name | POC | <year-2> Total | <year-1> Total | <year-1> Margin |
    Q1 Total | Q1 Margin | Q2 Total | Q2 Margin | Q3 Total | Q3 Margin |
    Q4 Total | Q4 Margin | Total | Margin | Comments
      \_______________________/  \_______________________________________/
            prior years                        current year

Each quarter gets two columns -- Total, then Margin -- under a merged
"Q1"/"Q2"/"Q3"/"Q4" group header, with a third header row underneath the
year band for those two sub-labels. Quarter Margin is populated straight
from `GroupSummary.quarter_margins` (computed in aggregator.py using the
exact same monthly-margin figures and month-to-quarter mapping already
used for the yearly Margin column -- see aggregator.py's docstring on
`quarter_margins` for exactly where that lives). Quarter Total is
unchanged: still `GroupSummary.quarters`, the existing revenue
aggregation. Only the presentation changed, not either calculation.

Formulas (not hardcoded values) are used wherever the value is derived
from cells written on the SAME sheet -- current-year Total, current-year
Margin, and every subtotal row -- exactly matching the convention used
in the manually built workbook. Because the four quarter Total (and
Margin) sub-columns are no longer contiguous (each quarter's Total and
Margin columns sit next to each other, not stacked), both the
current-year Total and current-year Margin formulas sum their four
quarter cells by explicit cell reference (`=SUM(F5,H5,J5,L5)`) rather
than a contiguous range, which would otherwise pull the wrong column's
values in. Figures that come from a *different* workbook/sheet
(quarters, margin, prior-year totals) are necessarily written as
computed values, the same way the original human-built Summary does it.

The header-writing code below never hardcodes which of "total"/"margin"
comes first within a quarter's pair -- it always resolves the pair's
own leftmost/rightmost column via min()/max() of
`self.quarter_cols[q].values()`, so `_plan_columns()`'s
`{"total": idx, "margin": idx + 1}` ordering is the single place that
decides the on-sheet order; nothing downstream needs to agree with it by
coincidence.

Every populated cell gets a thin black border (Feature 1) and every
monetary column gets a proper Excel currency number format with a real
"$" symbol, not text (Feature 4) -- see config.CURRENCY_FORMAT and
config.BORDER_COLOR. A second, currently-empty worksheet named
"<year> Actual & Forecast" is also created (Feature 5); it carries no
business logic yet.
"""
from __future__ import annotations

import calendar
import re
import zipfile
import xml.etree.ElementTree as ET
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config
from aggregator import GroupSummary

THIN_BLACK_BORDER = Border(
    left=Side(style="thin", color=config.BORDER_COLOR),
    right=Side(style="thin", color=config.BORDER_COLOR),
    top=Side(style="thin", color=config.BORDER_COLOR),
    bottom=Side(style="thin", color=config.BORDER_COLOR),
)


def _sum_formula_cached_value(ws: Worksheet, formula: str) -> float:
    """Evaluate a `=SUM(...)` formula this module just wrote, purely by
    reading the already-written sibling cells it references on the same
    sheet -- no business logic, just arithmetic on numbers already
    sitting on the sheet. Supports exactly the two shapes this module
    ever writes: a comma-separated list of cell refs
    (``=SUM(F5,H5,J5,L5)``) and a single contiguous range
    (``=SUM(C6:C33)``). This never invents or recomputes a business
    figure -- it can only ever reproduce the same number the formula
    itself would show once Excel (or any other viewer) evaluates it.

    A referenced cell may itself hold another ``=SUM(...)`` formula
    (e.g. a subtotal row summing a column of per-group Total cells,
    each of which is itself a formula over that row's own quarter
    cells) -- resolved by recursing into the same evaluator, since
    openpyxl's in-memory ``Cell.value`` for a formula cell is the
    formula text itself, not a number. This always terminates: every
    formula this module writes only ever references cells written
    earlier in the same top-down build, bottoming out at a literal
    numeric cell.
    """
    inner = formula[1:] if formula.startswith("=") else formula
    match = re.fullmatch(r"SUM\((.+)\)", inner, re.IGNORECASE)
    if not match:
        raise ValueError(f"Unsupported formula shape for cached-value evaluation: {formula!r}")
    body = match.group(1)

    def _resolve(cell) -> float:
        value = cell.value
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, str) and value.startswith("="):
            return _sum_formula_cached_value(ws, value)
        return 0.0

    total = 0.0
    if "," in body:
        for ref in body.split(","):
            total += _resolve(ws[ref.strip()])
    else:
        start_ref, end_ref = body.split(":")
        for row in ws[f"{start_ref.strip()}:{end_ref.strip()}"]:
            for cell in row:
                total += _resolve(cell)
    return total


def _format_cached_number(value: float) -> str:
    """Render a cached numeric value for the `<v>` element the same way
    a whole-number float should look (``"0"``, not ``"0.0"``) while
    avoiding float-arithmetic artifacts for fractional values -- Excel's
    own numeric XML parser accepts either form identically, so this only
    needs to be a valid, clean numeric literal, not a byte-for-byte match
    of openpyxl's own (undocumented) formatting."""
    rounded = round(value, 6)
    if rounded == int(rounded):
        return str(int(rounded))
    return repr(rounded)


_SML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
ET.register_namespace("", _SML_NS)  # write plain <c>, not <ns0:c>, on re-serialization


def _inject_cached_formula_values(path, entries: List[Tuple[str, str, float]]) -> None:
    """After ``wb.save(path)``, patch in a cached numeric result
    alongside each live formula cell tracked in ``entries`` (a list of
    ``(sheet_title, cell_coordinate, value)``), so every formula cell
    this module wrote displays its correct number the instant the
    workbook is opened -- in Excel's Normal Open, Protected View, a
    Quick Look/preview pane, or any other viewer that renders a cell's
    last-saved cached value without running a calculation engine --
    rather than depending on that viewer to recalculate on load. This
    only fills in the same (currently-empty) ``<v>`` element openpyxl
    already writes next to every formula's ``<f>`` element; it does not
    touch the formula itself, add a cell, or change any other part of
    the file.

    Each target sheet's XML is parsed into a real element tree
    (``xml.etree.ElementTree``) and edited via the tree API -- finding
    each ``<c>`` by its ``r`` attribute and setting its ``<v>`` child's
    text -- then re-serialized as a whole, rather than by locating and
    splicing substrings by hand. Every edit therefore starts from, and
    produces, a parsed (by definition well-formed) tree: there is no
    way for this to emit a mismatched or duplicated tag, which a naive
    string/regex splice of raw XML text does not equally guarantee (an
    earlier version of this function did exactly that, and it produced
    invalid OOXML -- e.g. a stray, unmatched `</v>` -- whenever a `<v>`
    element wasn't already empty; Excel's repair dialog on
    `/xl/worksheets/sheet1.xml` and `/xl/worksheets/sheet2.xml` was that
    defect, not anything about which cells get a cached value).
    """
    if not entries:
        return

    by_sheet: Dict[str, List[Tuple[str, float]]] = {}
    for sheet_title, coordinate, value in entries:
        by_sheet.setdefault(sheet_title, []).append((coordinate, value))

    with zipfile.ZipFile(path, "r") as zin:
        workbook_xml = zin.read("xl/workbook.xml").decode("utf-8")
        rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        original_items = {item.filename: zin.read(item.filename) for item in zin.infolist()}
        infolist = zin.infolist()

    # sheet title -> r:id (from <sheet name="..." r:id="..."/> in workbook.xml),
    # attributes extracted independently for the same reason as rid_to_target
    # below.
    title_to_rid: Dict[str, str] = {}
    for sheet_tag in re.findall(r"<sheet\b[^>]*/>", workbook_xml):
        name_match = re.search(r'\bname="([^"]*)"', sheet_tag)
        rid_match = re.search(r'\br:id="([^"]*)"', sheet_tag)
        if name_match and rid_match:
            title_to_rid[name_match.group(1)] = rid_match.group(1)
    # r:id -> zip-internal worksheet path (from workbook.xml.rels). Each
    # <Relationship .../> tag's Id and Target attributes are extracted
    # independently rather than assuming a fixed attribute order, since
    # relying on textual order (Target-before-Id vs Id-before-Target)
    # is exactly what silently broke this the first time.
    rid_to_target: Dict[str, str] = {}
    for rel_tag in re.findall(r"<Relationship\b[^>]*/>", rels_xml):
        rel_id_match = re.search(r'\bId="([^"]*)"', rel_tag)
        target_match = re.search(r'\bTarget="([^"]*)"', rel_tag)
        if rel_id_match and target_match:
            rid_to_target[rel_id_match.group(1)] = target_match.group(1)

    def _resolve_sheet_path(sheet_title: str) -> Optional[str]:
        rid = title_to_rid.get(sheet_title.replace("&", "&amp;"))
        target = rid_to_target.get(rid) if rid else None
        if not target:
            return None
        target = target.lstrip("/")
        return target if target.startswith("xl/") else f"xl/{target}"

    modified_items = dict(original_items)
    for sheet_title, cell_entries in by_sheet.items():
        sheet_path = _resolve_sheet_path(sheet_title)
        if sheet_path is None or sheet_path not in modified_items:
            continue  # defensive: never fail the whole save over a cosmetic patch
        patched_xml = _patch_sheet_cached_values(modified_items[sheet_path], cell_entries)
        if patched_xml is not None:
            modified_items[sheet_path] = patched_xml

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in infolist:
            zout.writestr(item, modified_items[item.filename])


def _patch_sheet_cached_values(sheet_xml: bytes, cell_entries: List[Tuple[str, float]]) -> Optional[bytes]:
    """Parse one worksheet's XML, set the cached ``<v>`` value for each
    ``(coordinate, value)`` in ``cell_entries``, and return the
    re-serialized bytes -- or ``None`` (leaving that sheet's original
    bytes untouched) if the XML can't be parsed, since this is a
    cosmetic "show the number immediately" patch layered on top of an
    already fully-correct workbook and must never be able to turn a
    successful generation into a failed or corrupted one.
    """
    try:
        root = ET.fromstring(sheet_xml)
    except ET.ParseError:
        return None

    cell_by_coordinate = {c.get("r"): c for c in root.iter(f"{{{_SML_NS}}}c")}
    for coordinate, value in cell_entries:
        cell = cell_by_coordinate.get(coordinate)
        if cell is None:
            continue  # defensive: this specific cell wasn't found -- skip, don't fail
        value_elem = cell.find(f"{{{_SML_NS}}}v")
        if value_elem is None:
            value_elem = ET.SubElement(cell, f"{{{_SML_NS}}}v")
        value_elem.text = _format_cached_number(value)

    # `xml_declaration=False` matches openpyxl's own convention for this
    # file (no `<?xml ... ?>` prologue on any part it writes) -- purely
    # cosmetic; Excel accepts a part with or without one.
    return ET.tostring(root, encoding="UTF-8", xml_declaration=False)


class SummaryWriter:
    def __init__(self, target_year: int, prior_years: List[int], years_with_margin: List[int]):
        self.target_year = target_year
        self.prior_years = sorted(prior_years)
        self.years_with_margin = set(years_with_margin)
        self._plan_columns()
        # (sheet_title, cell_coordinate, cached_value) for every live
        # formula cell written by this instance -- see
        # `patch_cached_formula_values`.
        self._formula_cache: List[Tuple[str, str, float]] = []

    # ------------------------------------------------------------------
    def _plan_columns(self) -> None:
        self.col_name = 1
        self.col_poc = 2
        idx = 3
        self.prior_year_cols: Dict[int, Dict[str, int]] = {}
        for year in self.prior_years:
            entry = {"total": idx}
            idx += 1
            if year in self.years_with_margin:
                entry["margin"] = idx
                idx += 1
            self.prior_year_cols[year] = entry

        # Each quarter occupies TWO columns -- Total, then Margin (per
        # the explicit ordering requirement: "Total, Margin, NOT
        # Margin, Total").
        self.quarter_cols: Dict[str, Dict[str, int]] = {}
        for q in config.QUARTER_ORDER:
            self.quarter_cols[q] = {"total": idx, "margin": idx + 1}
            idx += 2

        self.col_current_total = idx
        idx += 1
        self.col_current_margin = idx
        idx += 1
        self.col_comments = idx
        self.last_col = idx

        # Every numeric/monetary column: used for subtotal formulas and
        # for the currency number format (Feature 4).
        self.numeric_cols: List[int] = (
            [c["total"] for c in self.prior_year_cols.values()]
            + [c["margin"] for c in self.prior_year_cols.values() if "margin" in c]
            + [c[part] for c in self.quarter_cols.values() for part in ("total", "margin")]
            + [self.col_current_total, self.col_current_margin]
        )

    # ------------------------------------------------------------------
    def patch_cached_formula_values(self, path) -> None:
        """Call this once, right after ``wb.save(path)`` (see
        ``main.py``/``gui/runner.py``), to fill in a cached numeric
        result for every live formula this writer wrote (Worksheet 1's
        per-group Total/Margin and every subtotal row, plus Worksheet
        2's equivalents). openpyxl never writes a cached result for a
        formula cell -- Excel itself normally fills that in the first
        time a human opens the file, recalculates, and saves it again.
        Until then, some viewers (Excel's Protected View, preview
        panes, and other lightweight renderers that show a cell's
        last-saved value without running a calculation engine) show
        those cells blank, which is exactly the "totals only appear
        after I click into the cell" symptom this fixes. The formulas
        themselves are completely untouched.
        """
        _inject_cached_formula_values(path, self._formula_cache)

    def build(
        self,
        sections: List[Tuple[config.OutputSection, List[GroupSummary]]],
        monthly_sections: Optional[List[Tuple[config.OutputSection, List["MonthlyGroupSummary"]]]] = None,
        month_roles: Optional[Dict[int, str]] = None,
    ) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = str(self.target_year)

        self._content_rows: List[int] = [1, 2, 3]  # header rows always bordered
        self._write_headers(ws)
        current_row = 4

        for section, groups in sections:
            if section.heading:
                self._write_banner_row(ws, current_row, section.heading, fill=config.SECTION_FILL)
                current_row += 1
                current_row += section.blank_rows_after_heading

            self._write_banner_row(ws, current_row, section.title, fill=config.SUBHEADING_FILL)
            current_row += 1
            current_row += section.blank_rows_after_title

            data_start = current_row
            for group in groups:
                self._write_group_row(ws, current_row, group, section.show_poc)
                current_row += 1
            data_end = current_row - 1

            current_row += section.blank_rows_after_data
            if data_end >= data_start:
                self._write_subtotal_row(ws, current_row, section.subtotal_label, data_start, data_end)
            else:
                self._write_banner_row(ws, current_row, section.subtotal_label, fill=config.SUBTOTAL_FILL)
            current_row += 1
            current_row += section.blank_rows_after_subtotal

        self._apply_sheet_formatting(ws, current_row)
        self._apply_borders(ws)  # Feature 1 -- after all content/formatting is in place

        # Feature 5 -- second worksheet, dynamically named.
        second_sheet_name = f"{self.target_year} Actual & Forecast"
        if monthly_sections is not None and month_roles:
            self._build_monthly_sheet(wb, second_sheet_name, monthly_sections, month_roles)
        else:
            # Defensive fallback only -- both current call sites
            # (main.py, gui/runner.py) always supply monthly_sections.
            # Kept so a hypothetical future caller that doesn't pass the
            # new arguments still gets a validly-named, non-broken sheet
            # instead of a hard failure.
            ws2 = wb.create_sheet(title=second_sheet_name)
            placeholder = ws2.cell(
                row=1,
                column=1,
                value=(
                    f"{self.target_year} Actual & Forecast -- detailed content will be "
                    "added in a future update."
                ),
            )
            placeholder.font = Font(name=config.FONT_NAME, size=config.FONT_SIZE, italic=True, color="FF808080")
            ws2.sheet_view.showGridLines = False
            ws2.column_dimensions["A"].width = 60

        return wb

    # ------------------------------------------------------------------
    def _write_headers(self, ws: Worksheet) -> None:
        bold_center = Font(name=config.FONT_NAME, size=config.FONT_SIZE, bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Row 1: year banners. Prior years merge over their own
        # column(s) exactly as before; the current year's merge now
        # spans the whole widened block (all 8 quarter sub-columns plus
        # the final Total/Margin columns).
        for year, cols in self.prior_year_cols.items():
            start = cols["total"]
            end = cols.get("margin", cols["total"])
            if end > start:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            cell = ws.cell(row=1, column=start, value=year)
            cell.font = bold_center
            cell.alignment = center

        cur_start = min(self.quarter_cols["Q1"].values())
        cur_end = self.col_current_margin
        ws.merge_cells(start_row=1, start_column=cur_start, end_row=1, end_column=cur_end)
        cell = ws.cell(row=1, column=cur_start, value=self.target_year)
        cell.font = bold_center
        cell.alignment = center

        # Row 2 + 3: columns with no sub-grouping (Name, POC, each prior
        # year's Total/Margin, the final yearly Total/Margin, Comments)
        # get their label on row 2, merged vertically down through row 3
        # so it reads as a single label beneath the year band -- the
        # same visual convention the original 2-row header used, just
        # one row taller. Quarter columns instead get "Q1".."Q4" merged
        # horizontally across their 2 sub-columns on row 2, with the
        # actual "Margin"/"Total" sub-labels on row 3 underneath.
        vertical_label_cols: Dict[int, str] = {
            self.col_name: "Name",
            self.col_poc: "POC",
            self.col_current_total: "Total",
            self.col_current_margin: "Margin",
            self.col_comments: "Comments",
        }
        for year, cols in self.prior_year_cols.items():
            vertical_label_cols[cols["total"]] = "Total"
            if "margin" in cols:
                vertical_label_cols[cols["margin"]] = "Margin"

        for col, label in vertical_label_cols.items():
            ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
            cell = ws.cell(row=2, column=col, value=label)
            cell.font = bold_center
            cell.alignment = center

        for q, cols in self.quarter_cols.items():
            start, end_col = min(cols.values()), max(cols.values())
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end_col)
            q_cell = ws.cell(row=2, column=start, value=q)
            q_cell.font = bold_center
            q_cell.alignment = center

            total_cell = ws.cell(row=3, column=cols["total"], value="Total")
            total_cell.font = bold_center
            total_cell.alignment = center
            margin_cell = ws.cell(row=3, column=cols["margin"], value="Margin")
            margin_cell.font = bold_center
            margin_cell.alignment = center

        for row in (1, 2, 3):
            for col in range(1, self.last_col + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=config.HEADER_FILL)

        # Total/Margin columns get the NK-workbook-sourced header color,
        # overriding the general header fill for just these two columns.
        for row in (1, 2, 3):
            for col in (self.col_current_total, self.col_current_margin):
                ws.cell(row=row, column=col).fill = PatternFill(
                    "solid", fgColor=config.TOTAL_MARGIN_HEADER_FILL
                )

    def _write_sum_formula(self, ws: Worksheet, row: int, column: int, formula: str):
        """Write one ``=SUM(...)`` formula cell and record the number it
        will evaluate to (computed by reading the sibling cells it
        references, already written by the time this runs) so
        ``patch_cached_formula_values`` can give it a cached result
        after the workbook is saved. Every formula this module writes
        goes through this one helper -- see its module-level docstring
        note on formulas for why a formula (not a value) is used here at
        all; this only changes how it additionally gets a cached result.
        """
        cell = ws.cell(row=row, column=column, value=formula)
        self._formula_cache.append((ws.title, cell.coordinate, _sum_formula_cached_value(ws, formula)))
        return cell

    def _write_banner_row(self, ws: Worksheet, row: int, text: str, fill: Optional[str] = None) -> None:
        self._content_rows.append(row)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=self.last_col)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(name=config.FONT_NAME, size=config.FONT_SIZE, bold=True)
        if fill:
            for col in range(1, self.last_col + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=fill)

    def _write_group_row(self, ws: Worksheet, row: int, group: GroupSummary, show_poc: bool) -> None:
        self._content_rows.append(row)
        ws.cell(row=row, column=self.col_name, value=group.group_name)
        if show_poc and group.poc:
            ws.cell(row=row, column=self.col_poc, value=group.poc)

        total_fill = PatternFill("solid", fgColor=config.TOTAL_DATA_FILL)
        margin_fill = PatternFill("solid", fgColor=config.MARGIN_DATA_FILL)

        for year, cols in self.prior_year_cols.items():
            total, margin = group.historical.get(year, (None, None))
            if total is not None:
                ws.cell(row=row, column=cols["total"], value=round(total, 2))
            ws.cell(row=row, column=cols["total"]).fill = total_fill
            if "margin" in cols:
                if margin is not None:
                    ws.cell(row=row, column=cols["margin"], value=round(margin, 2))
                ws.cell(row=row, column=cols["margin"]).fill = margin_fill

        total_col_refs: List[str] = []
        margin_col_refs: List[str] = []
        for q, cols in self.quarter_cols.items():
            ws.cell(row=row, column=cols["margin"], value=group.quarter_margins.get(q, 0.0))
            ws.cell(row=row, column=cols["total"], value=group.quarters.get(q, 0.0))
            ws.cell(row=row, column=cols["margin"]).fill = margin_fill
            ws.cell(row=row, column=cols["total"]).fill = total_fill
            total_col_refs.append(f"{get_column_letter(cols['total'])}{row}")
            margin_col_refs.append(f"{get_column_letter(cols['margin'])}{row}")

        # Final yearly Total = sum of the four quarter Total sub-columns.
        # They're no longer contiguous (each quarter's Margin column
        # sits between them), so this sums them by explicit cell
        # reference rather than a contiguous range -- a mechanical
        # adjustment to match the new layout, not a methodology change;
        # the value is identical to summing Q1..Q4 revenue as before.
        self._write_sum_formula(ws, row, self.col_current_total, f"=SUM({','.join(total_col_refs)})")
        # Final yearly Margin: now also a live formula, summing the same
        # four quarter Margin cells already written above -- mirroring
        # the Total column exactly. This produces the identical number
        # `group.total_margin` already holds (both are sums of the same
        # underlying monthly margin figures; the quarter-margin-sums-to-
        # yearly-margin identity is already covered by
        # tests/compare_with_manual.py), so no calculation changes --
        # only how the value is expressed on the sheet.
        self._write_sum_formula(ws, row, self.col_current_margin, f"=SUM({','.join(margin_col_refs)})")
        ws.cell(row=row, column=self.col_current_total).fill = total_fill
        ws.cell(row=row, column=self.col_current_margin).fill = margin_fill

        if group.comment:  # Rule 6: comment blanks stay blank
            c = ws.cell(row=row, column=self.col_comments, value=group.comment)
            # `wrap_text` deliberately left off here, matching the
            # manually-built template's own Comments-column cells
            # (verified: `vertical="top"` only, `wrapText` unset) --
            # turning it on is what makes Excel auto-expand a row's
            # height to fit a long comment, so a row's height would
            # vary with comment length instead of staying uniform.
            c.alignment = Alignment(vertical="top")

    def _write_subtotal_row(self, ws: Worksheet, row: int, label: str, data_start: int, data_end: int) -> None:
        self._content_rows.append(row)
        cell = ws.cell(row=row, column=self.col_name, value=label)
        cell.font = Font(name=config.FONT_NAME, size=config.FONT_SIZE, bold=True)
        for col in self.numeric_cols:
            letter = get_column_letter(col)
            formula_cell = self._write_sum_formula(ws, row, col, f"=SUM({letter}{data_start}:{letter}{data_end})")
            formula_cell.font = Font(name=config.FONT_NAME, size=config.FONT_SIZE, bold=True)
        for col in range(1, self.last_col + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=config.SUBTOTAL_FILL)

    # ------------------------------------------------------------------
    def _apply_sheet_formatting(self, ws: Worksheet, last_row: int) -> None:
        base_font = Font(name=config.FONT_NAME, size=config.FONT_SIZE)
        for row in ws.iter_rows(min_row=1, max_row=max(last_row, 3), min_col=1, max_col=self.last_col):
            for cell in row:
                if cell.font is None or cell.font.name is None:
                    cell.font = base_font

        # Feature 4: proper Excel currency format (a real "$", not text)
        # on every monetary column; values themselves stay numeric.
        for col in self.numeric_cols:
            for row in range(4, last_row + 1):
                ws.cell(row=row, column=col).number_format = config.CURRENCY_FORMAT

        ws.column_dimensions[get_column_letter(self.col_name)].width = config.NAME_COLUMN_WIDTH
        ws.column_dimensions[get_column_letter(self.col_poc)].width = config.POC_COLUMN_WIDTH
        for col in self.numeric_cols:
            ws.column_dimensions[get_column_letter(col)].width = config.NUMBER_COLUMN_WIDTH
        ws.column_dimensions[get_column_letter(self.col_comments)].width = config.COMMENTS_COLUMN_WIDTH

        ws.freeze_panes = "A4"
        ws.sheet_view.showGridLines = False

    def _apply_borders(self, ws: Worksheet) -> None:
        """Feature 1: thin black border on every populated cell -- every
        header row, every data row, every section title/subtotal/total
        row, including the Comments column. Intentionally skips the
        blank spacer rows between sections (config.OutputSection's
        blank_rows_after_* rows) so they stay purely as visual spacing.
        Only touches `.border`; fills, fonts, alignment, merges and
        number formats already applied elsewhere are left exactly as
        they are.
        """
        for row in self._content_rows:
            for col in range(1, self.last_col + 1):
                ws.cell(row=row, column=col).border = THIN_BLACK_BORDER

    # ------------------------------------------------------------------
    # Worksheet 2: "<year> Actual & Forecast"
    # ------------------------------------------------------------------
    def _build_monthly_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        monthly_sections: List[Tuple[config.OutputSection, List["MonthlyGroupSummary"]]],
        month_roles: Dict[int, str],
    ) -> None:
        """Builds the month-by-month Actual/Forecast worksheet.

        Deliberately mirrors Worksheet 1's own visual language (the same
        merged-group-header convention, the same fills/fonts/borders/
        currency format constants from config.py) rather than
        introducing a second style, and reuses the exact same
        section/spacing model (`config.OutputSection`'s heading/title/
        subtotal/blank-row fields) that already drives Worksheet 1's
        layout -- so a change to a section's spacing in config.py
        affects both sheets consistently instead of needing to be kept
        in sync by hand.

        Column layout (fully dynamic -- driven entirely by whatever
        months and role labels `month_roles` actually contains, in
        calendar order):

            Name | POC | <Role> <Mon> | Margin | <Role> <Mon> | Margin | ...
            Total | Margin | Confidence | Comments

        Every group's Total and Margin are both live `=SUM(...)`
        formulas over that row's own monthly value/margin cells
        (mirroring Worksheet 1's formula convention for both columns);
        Confidence and Comments are copied from the already-validated
        `MonthlyGroupSummary` -- confidence from the sheet's own Renewal
        Confidence column, and Comments from the exact same
        already-matched value Worksheet 1 shows.
        """
        ws2 = wb.create_sheet(title=sheet_name)
        months = sorted(month_roles.keys())

        # -- Column plan --------------------------------------------
        col_name = 1
        col_poc = 2
        month_cols: Dict[int, Dict[str, int]] = {}
        idx = 3
        for m in months:
            month_cols[m] = {"value": idx, "margin": idx + 1}
            idx += 2
        col_total = idx
        idx += 1
        col_margin = idx
        idx += 1
        col_confidence = idx
        idx += 1
        col_comments = idx
        last_col = idx

        numeric_cols: List[int] = (
            [c[part] for c in month_cols.values() for part in ("value", "margin")]
            + [col_total, col_margin]
        )

        content_rows: List[int] = [1, 2]

        # -- Headers ---------------------------------------------------
        bold_center = Font(name=config.FONT_NAME, size=config.FONT_SIZE, bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        vertical_labels: Dict[int, str] = {
            col_name: "Name",
            col_poc: "POC",
            col_total: "Total",
            col_margin: "Margin",
            col_confidence: "Confidence",
            col_comments: "Comments",
        }
        for col, label in vertical_labels.items():
            ws2.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            cell = ws2.cell(row=1, column=col, value=label)
            cell.font = bold_center
            cell.alignment = center

        for m, cols in month_cols.items():
            start, end = min(cols.values()), max(cols.values())
            ws2.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            role_cell = ws2.cell(row=1, column=start, value=month_roles.get(m, ""))
            role_cell.font = bold_center
            role_cell.alignment = center

            month_abbr = calendar.month_abbr[m] if 1 <= m <= 12 else str(m)
            value_cell = ws2.cell(row=2, column=cols["value"], value=month_abbr)
            value_cell.font = bold_center
            value_cell.alignment = center
            margin_cell = ws2.cell(row=2, column=cols["margin"], value="Margin")
            margin_cell.font = bold_center
            margin_cell.alignment = center

        for row in (1, 2):
            for col in range(1, last_col + 1):
                ws2.cell(row=row, column=col).fill = PatternFill("solid", fgColor=config.HEADER_FILL)

        # Total/Margin columns get the NK-workbook-sourced header color,
        # overriding the general header fill for just these two columns
        # -- same treatment as Worksheet 1's final Total/Margin columns.
        for row in (1, 2):
            for col in (col_total, col_margin):
                ws2.cell(row=row, column=col).fill = PatternFill(
                    "solid", fgColor=config.TOTAL_MARGIN_HEADER_FILL
                )

        # -- Body: same section/spacing model as Worksheet 1 ------------
        current_row = 3
        for section, groups in monthly_sections:
            if section.heading:
                self._write_plain_banner(ws2, current_row, section.heading, last_col, content_rows, config.SECTION_FILL)
                current_row += 1
                current_row += section.blank_rows_after_heading

            self._write_plain_banner(ws2, current_row, section.title, last_col, content_rows, config.SUBHEADING_FILL)
            current_row += 1
            current_row += section.blank_rows_after_title

            data_start = current_row
            total_fill = PatternFill("solid", fgColor=config.TOTAL_DATA_FILL)
            margin_fill = PatternFill("solid", fgColor=config.MARGIN_DATA_FILL)
            for g in groups:
                content_rows.append(current_row)
                ws2.cell(row=current_row, column=col_name, value=g.group_name)
                # POC is always shown on Worksheet 2 when available, regardless
                # of Worksheet 1's per-section `show_poc` flag. That flag
                # governs Worksheet 1's own presentation (e.g. Track 1 hides
                # POC there); Worksheet 2 is a different report and showing
                # POC for every row matches both the NK workbook and the
                # business-provided screenshot, which populate POC for every
                # Track 1 row. See audit finding #1.
                if g.poc:
                    ws2.cell(row=current_row, column=col_poc, value=g.poc)

                value_refs: List[str] = []
                margin_refs: List[str] = []
                for m, cols in month_cols.items():
                    ws2.cell(row=current_row, column=cols["value"], value=g.monthly_revenue.get(m, 0.0))
                    ws2.cell(row=current_row, column=cols["margin"], value=g.monthly_margin.get(m, 0.0))
                    ws2.cell(row=current_row, column=cols["value"]).fill = total_fill
                    ws2.cell(row=current_row, column=cols["margin"]).fill = margin_fill
                    value_refs.append(f"{get_column_letter(cols['value'])}{current_row}")
                    margin_refs.append(f"{get_column_letter(cols['margin'])}{current_row}")

                # Total: live formula over this row's own monthly value
                # cells (mirrors Worksheet 1's Total-formula convention).
                if value_refs:
                    self._write_sum_formula(ws2, current_row, col_total, f"=SUM({','.join(value_refs)})")
                else:
                    ws2.cell(row=current_row, column=col_total, value=0)
                # Margin: also a live formula now, summing this row's own
                # monthly Margin cells written just above -- mirroring
                # the Total column exactly. Produces the identical
                # number `g.total_margin` already holds (both are sums
                # of the same underlying monthly margin figures), so no
                # calculation changes -- only how the value is expressed.
                if margin_refs:
                    self._write_sum_formula(ws2, current_row, col_margin, f"=SUM({','.join(margin_refs)})")
                else:
                    ws2.cell(row=current_row, column=col_margin, value=0)
                ws2.cell(row=current_row, column=col_total).fill = total_fill
                ws2.cell(row=current_row, column=col_margin).fill = margin_fill

                if g.confidence:
                    ws2.cell(row=current_row, column=col_confidence, value=g.confidence)
                if g.comment:
                    c = ws2.cell(row=current_row, column=col_comments, value=g.comment)
                    # Same reasoning as Worksheet 1's comment cells: no
                    # `wrap_text`, matching the template, so row height
                    # doesn't vary with comment length.
                    c.alignment = Alignment(vertical="top")
                current_row += 1
            data_end = current_row - 1

            current_row += section.blank_rows_after_data
            if data_end >= data_start:
                content_rows.append(current_row)
                cell = ws2.cell(row=current_row, column=col_name, value=section.subtotal_label)
                cell.font = Font(name=config.FONT_NAME, size=config.FONT_SIZE, bold=True)
                for col in numeric_cols:
                    letter = get_column_letter(col)
                    fcell = self._write_sum_formula(
                        ws2, current_row, col, f"=SUM({letter}{data_start}:{letter}{data_end})",
                    )
                    fcell.font = Font(name=config.FONT_NAME, size=config.FONT_SIZE, bold=True)
                for col in range(1, last_col + 1):
                    ws2.cell(row=current_row, column=col).fill = PatternFill("solid", fgColor=config.SUBTOTAL_FILL)
            else:
                self._write_plain_banner(ws2, current_row, section.subtotal_label, last_col, content_rows, config.SUBTOTAL_FILL)
            current_row += 1
            current_row += section.blank_rows_after_subtotal

        last_row = current_row

        # -- Sheet-wide formatting (mirrors _apply_sheet_formatting) ----
        base_font = Font(name=config.FONT_NAME, size=config.FONT_SIZE)
        for row in ws2.iter_rows(min_row=1, max_row=max(last_row, 2), min_col=1, max_col=last_col):
            for cell in row:
                if cell.font is None or cell.font.name is None:
                    cell.font = base_font

        for col in numeric_cols:
            for row in range(3, last_row + 1):
                ws2.cell(row=row, column=col).number_format = config.CURRENCY_FORMAT

        ws2.column_dimensions[get_column_letter(col_name)].width = config.NAME_COLUMN_WIDTH
        ws2.column_dimensions[get_column_letter(col_poc)].width = config.POC_COLUMN_WIDTH
        for col in numeric_cols:
            ws2.column_dimensions[get_column_letter(col)].width = config.NUMBER_COLUMN_WIDTH
        ws2.column_dimensions[get_column_letter(col_confidence)].width = config.CONFIDENCE_COLUMN_WIDTH
        ws2.column_dimensions[get_column_letter(col_comments)].width = config.COMMENTS_COLUMN_WIDTH

        ws2.freeze_panes = "A3"
        ws2.sheet_view.showGridLines = False

        # -- Borders (Feature 1, mirrored) ------------------------------
        for row in content_rows:
            for col in range(1, last_col + 1):
                ws2.cell(row=row, column=col).border = THIN_BLACK_BORDER

    @staticmethod
    def _write_plain_banner(
        ws: Worksheet, row: int, text: str, last_col: int, content_rows: List[int], fill: str
    ) -> None:
        """Same visual behaviour as `_write_banner_row`, parameterised
        for an arbitrary worksheet/column-count/content-row-list instead
        of `self`'s Worksheet-1-specific `last_col`/`_content_rows`, so
        Worksheet 1's own `_write_banner_row` never needs to change.
        """
        content_rows.append(row)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(name=config.FONT_NAME, size=config.FONT_SIZE, bold=True)
        for col in range(1, last_col + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=fill)
