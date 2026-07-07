"""
tests/compare_with_manual.py
=============================
Automated regression test: regenerates the Summary workbook from the
fixture Master workbook and compares it against the golden,
manually-produced Summary workbook (which predates the Q1-Q4
Margin/Total grouped-column layout).

Because the output *layout* changed (each quarter is now two columns,
Margin then Total, under a merged "Q1".."Q4" header, with a 3-row header
instead of 2 -- see summary_writer.py), this test no longer assumes
fixed cell coordinates for either file. Both the golden file's header
(2 rows) and the generated file's header (3 rows) are parsed dynamically
by reading their own label text, then matched up conceptually:

    golden "Q1" (single revenue column)  <->  generated "Q1 Total"
    golden "Total" (final, yearly)        <->  generated final "Total"
    golden "Margin" (final, yearly)       <->  generated final "Margin"
    (golden has no per-quarter Margin column at all -- see the
     dedicated internal-consistency check below for how that's verified
     instead)

Rows are matched between the two files by the text in column A (Name /
section title / subtotal label), not by row number, so this test does
not depend on the two files' header heights lining up by coincidence.

Two kinds of check run:
1. Cross-file: every quantity that existed in the OLD layout (Name, POC,
   prior-year Total/Margin, each quarter's revenue Total, the final
   yearly Total/Margin, Comments) must still match the golden file
   exactly wherever a matching row exists in both.
2. New-layout self-consistency: for every data row, the four new
   quarterly Margin sub-columns must sum to that row's final yearly
   Margin (both are built from the exact same monthly-margin figures in
   aggregator.py, so they must agree) -- this is the only meaningful way
   to validate the new Quarter Margin columns, since no golden reference
   for them exists.

This test is intentionally dependency-free (no LibreOffice / no pytest
required): a tiny built-in evaluator resolves this project's own
`=SUM(...)` formulas, both the contiguous-range form used by subtotal
rows (`=SUM(F5:F10)`) and the explicit-cell-list form used by each data
row's yearly Total (`=SUM(G7,I7,K7,M7)`, needed because the four quarter
Total sub-columns are no longer contiguous).
"""
from __future__ import annotations

import re
import sys
import tempfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, range_boundaries

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import main as sfae_main  # noqa: E402

FIXTURE_MASTER_GLOB = "*.xlsx"
GOLDEN_SUMMARY = Path(__file__).resolve().parent / "fixtures" / "manual_summary_2026.xlsx"
GOLDEN_SHEET = "2026"
NUMERIC_TOLERANCE = 0.015  # ignore sub-₹0.01 floating point noise (float64 can land at ~0.0100000002)


def _find_fixture_master() -> Path:
    fixtures_dir = Path(__file__).resolve().parent / "fixtures"
    candidates = sorted(fixtures_dir.glob(FIXTURE_MASTER_GLOB))
    candidates = [c for c in candidates if c.name != GOLDEN_SUMMARY.name]
    if candidates:
        return candidates[0]
    input_candidates = sorted((PROJECT_ROOT / "input").glob("*.xlsx"))
    input_candidates = [c for c in input_candidates if not c.name.startswith("~$")]
    if input_candidates:
        return input_candidates[0]
    raise FileNotFoundError(
        "No Master workbook found for testing. Place one at "
        f"{fixtures_dir}/ or {PROJECT_ROOT}/input/."
    )


def _norm(value) -> str:
    return re.sub(r"\s+", " ", str(value).strip().lower()) if value is not None else ""


# --------------------------------------------------------------------
# Formula evaluation -- handles both formula shapes this project writes
# --------------------------------------------------------------------
_SUM_RANGE_RE = re.compile(r"^=SUM\(([A-Z]+\d+):([A-Z]+\d+)\)$", re.IGNORECASE)
_SUM_LIST_RE = re.compile(r"^=SUM\(([A-Z]+\d+(?:\s*,\s*[A-Z]+\d+)+)\)$", re.IGNORECASE)
_CELL_REF_RE = re.compile(r"^([A-Z]+)(\d+)$", re.IGNORECASE)


def resolve_cell(ws: Worksheet, row: int, col: int, _depth: int = 0):
    """Return a cell's effective value, evaluating this project's own
    `=SUM(...)` formulas -- either the contiguous-range form
    (subtotal rows) or the explicit comma-separated cell-list form
    (each data row's yearly Total, summing the 4 non-contiguous quarter
    Total sub-columns) -- recursively, since a subtotal's range can
    itself contain formula cells."""
    if _depth > 5:
        return ws.cell(row, col).value
    value = ws.cell(row, col).value
    if not (isinstance(value, str) and value.startswith("=")):
        return value

    text = value.strip()

    m = _SUM_RANGE_RE.match(text)
    if m:
        min_col, min_row, max_col, max_row = range_boundaries(f"{m.group(1)}:{m.group(2)}")
        total, saw_number = 0.0, False
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                v = resolve_cell(ws, r, c, _depth + 1)
                if isinstance(v, (int, float)):
                    total += v
                    saw_number = True
        return round(total, 2) if saw_number else 0

    m = _SUM_LIST_RE.match(text)
    if m:
        total, saw_number = 0.0, False
        for ref in [p.strip() for p in m.group(1).split(",")]:
            cm = _CELL_REF_RE.match(ref)
            if not cm:
                continue
            c = 0
            for ch in cm.group(1).upper():
                c = c * 26 + (ord(ch) - ord("A") + 1)
            v = resolve_cell(ws, int(cm.group(2)), c, _depth + 1)
            if isinstance(v, (int, float)):
                total += v
                saw_number = True
        return round(total, 2) if saw_number else 0

    return value  # unrecognised formula shape - compare literally


def _values_match(a, b) -> bool:
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(a - b) <= NUMERIC_TOLERANCE
    if a is None:
        a = ""
    if b is None:
        b = ""
    if isinstance(a, str) and isinstance(b, str):
        return a.strip() == b.strip()
    return a == b


# --------------------------------------------------------------------
# Header parsing -- dynamic, by label text, for each of the two shapes
# --------------------------------------------------------------------
class SheetColumns:
    """Resolved column positions for one Summary sheet, regardless of
    which of the two header shapes (old 2-row / new 3-row) it uses."""

    def __init__(self) -> None:
        self.name: Optional[int] = None
        self.poc: Optional[int] = None
        self.prior_years: List[Dict[str, int]] = []  # [{"total": col, "margin": col_or_None}, ...] in order
        self.quarters: Dict[str, Dict[str, int]] = {}  # {"Q1": {"total": col, "margin": col_or_None}, ...}
        self.total: Optional[int] = None
        self.margin: Optional[int] = None
        self.comments: Optional[int] = None


def parse_golden_columns(ws: Worksheet) -> SheetColumns:
    """Old layout, single header-label row (row 2): Name, POC,
    <prior-year Total[/Margin]>..., Q1, Q2, Q3, Q4, Total, Margin,
    Comments."""
    cols = SheetColumns()
    max_col = ws.max_column
    labels = [_norm(ws.cell(2, c).value) for c in range(1, max_col + 1)]

    i = 0
    while i < len(labels):
        col = i + 1
        label = labels[i]
        if label == "name" and cols.name is None:
            cols.name = col
        elif label == "poc" and cols.poc is None:
            cols.poc = col
        elif label == "comments":
            cols.comments = col
        elif re.fullmatch(r"q[1-4]", label):
            cols.quarters[label.upper()] = {"total": col, "margin": None}
        elif label in ("total", "margin"):
            # Sequence-dependent: prior-year Total/Margin pairs come
            # before Q1; the single final Total/Margin pair comes after
            # Q4. Distinguish by whether all four quarters are resolved
            # yet.
            if len(cols.quarters) < 4:
                if label == "total":
                    cols.prior_years.append({"total": col, "margin": None})
                else:
                    if cols.prior_years:
                        cols.prior_years[-1]["margin"] = col
            else:
                if label == "total":
                    cols.total = col
                else:
                    cols.margin = col
        i += 1
    return cols


def parse_generated_columns(ws: Worksheet) -> SheetColumns:
    """New layout, 2 header-label rows (row 2 = group-level labels /
    Q1-Q4 merged headers, row 3 = Margin/Total sub-labels under each
    quarter)."""
    cols = SheetColumns()
    max_col = ws.max_column
    row2 = [_norm(ws.cell(2, c).value) for c in range(1, max_col + 1)]
    row3 = [_norm(ws.cell(3, c).value) for c in range(1, max_col + 1)]

    i = 0
    while i < len(row2):
        col = i + 1
        label = row2[i]
        if label == "name" and cols.name is None:
            cols.name = col
        elif label == "poc" and cols.poc is None:
            cols.poc = col
        elif label == "comments":
            cols.comments = col
        elif re.fullmatch(r"q[1-4]", label):
            # Anchor column of the merged "Qn" header, spanning 2
            # sub-columns. Which of the two is Total and which is
            # Margin is read from row 3's own text at each position,
            # never assumed by position -- the sheet may order them
            # either way (Total-then-Margin is the current convention;
            # this parser doesn't hardcode that).
            first_label = row3[i]
            second_label = row3[i + 1] if i + 1 < len(row3) else ""
            sub_labels = {first_label: col, second_label: col + 1}
            cols.quarters[label.upper()] = {
                "total": sub_labels.get("total"),
                "margin": sub_labels.get("margin"),
            }
        elif label in ("total", "margin"):
            if len(cols.quarters) < 4:
                if label == "total":
                    cols.prior_years.append({"total": col, "margin": None})
                else:
                    if cols.prior_years:
                        cols.prior_years[-1]["margin"] = col
            else:
                if label == "total":
                    cols.total = col
                else:
                    cols.margin = col
        i += 1
    return cols


def _row_labels(ws: Worksheet, name_col: int, start_row: int) -> Dict[str, int]:
    """Map normalised column-A-ish (the resolved Name column) text to
    its row number, first occurrence wins. Starts after the header rows
    so the header's own "Name" label can't match itself as a pseudo-row."""
    labels: Dict[str, int] = {}
    for r in range(start_row, ws.max_row + 1):
        val = ws.cell(r, name_col).value
        key = _norm(val)
        if key and key not in labels:
            labels[key] = r
    return labels


# --------------------------------------------------------------------
# Comparison
# --------------------------------------------------------------------
def compare_workbooks(generated_path: Path, golden_path: Path, golden_sheet: str) -> Tuple[list, list]:
    generated_wb = load_workbook(generated_path, data_only=False)
    golden_wb = load_workbook(golden_path, data_only=True)

    if golden_sheet not in golden_wb.sheetnames:
        raise ValueError(f"Golden workbook has no sheet '{golden_sheet}'.")
    generated_ws = generated_wb[golden_sheet] if golden_sheet in generated_wb.sheetnames else generated_wb.active
    golden_ws = golden_wb[golden_sheet]

    gcols = parse_golden_columns(golden_ws)
    ncols = parse_generated_columns(generated_ws)
    for label, cols in (("golden", gcols), ("generated", ncols)):
        if cols.name is None:
            raise ValueError(f"Could not locate a 'Name' column in the {label} sheet's header.")
        if len(cols.quarters) != 4:
            raise ValueError(f"Expected 4 quarter columns (Q1-Q4) in the {label} sheet, found {len(cols.quarters)}.")

    golden_rows = _row_labels(golden_ws, gcols.name, start_row=3)
    generated_rows = _row_labels(generated_ws, ncols.name, start_row=4)
    shared_labels = [lbl for lbl in golden_rows if lbl in generated_rows]

    diffs = []
    for label in shared_labels:
        gr, nr = golden_rows[label], generated_rows[label]

        def check(field: str, gcol: Optional[int], ncol: Optional[int]) -> None:
            if gcol is None or ncol is None:
                return
            gval = resolve_cell(golden_ws, gr, gcol)
            nval = resolve_cell(generated_ws, nr, ncol)
            if gval is None and nval is None:
                return
            if not _values_match(gval, nval):
                diffs.append(
                    {
                        "label": label,
                        "field": field,
                        "golden_coord": f"{get_column_letter(gcol)}{gr}",
                        "generated_coord": f"{get_column_letter(ncol)}{nr}",
                        "expected": gval,
                        "actual": nval,
                    }
                )

        check("Name", gcols.name, ncols.name)
        check("POC", gcols.poc, ncols.poc)
        for idx, (gpy, npy) in enumerate(zip(gcols.prior_years, ncols.prior_years)):
            check(f"PriorYear[{idx}].Total", gpy.get("total"), npy.get("total"))
            check(f"PriorYear[{idx}].Margin", gpy.get("margin"), npy.get("margin"))
        for q in ("Q1", "Q2", "Q3", "Q4"):
            # Golden's single quarter column is that quarter's revenue
            # Total -- compared against the new layout's "<Q> Total"
            # sub-column (Quarter Margin has no golden equivalent; see
            # the self-consistency check below instead).
            check(f"{q}.Total", gcols.quarters[q]["total"], ncols.quarters[q]["total"])
        check("Total", gcols.total, ncols.total)
        check("Margin", gcols.margin, ncols.margin)
        check("Comments", gcols.comments, ncols.comments)

    # Self-consistency check (no golden equivalent exists): each row's
    # 4 new quarterly Margin sub-columns must sum to that row's final
    # yearly Margin, since both come from the exact same monthly-margin
    # figures in aggregator.py.
    consistency_failures = []
    for label, nr in generated_rows.items():
        margin_cols = [ncols.quarters[q]["margin"] for q in ("Q1", "Q2", "Q3", "Q4")]
        margin_vals = [resolve_cell(generated_ws, nr, c) for c in margin_cols]
        if not any(isinstance(v, (int, float)) and v for v in margin_vals):
            continue  # blank/zero row (e.g. a section title row) - nothing to check
        quarter_sum = sum(v for v in margin_vals if isinstance(v, (int, float)))
        yearly_margin = resolve_cell(generated_ws, nr, ncols.margin) if ncols.margin else None
        if yearly_margin is None:
            continue
        if not _values_match(round(quarter_sum, 2), yearly_margin):
            consistency_failures.append(
                {
                    "label": label,
                    "row": nr,
                    "quarter_margin_sum": round(quarter_sum, 2),
                    "yearly_margin": yearly_margin,
                }
            )

    return diffs, consistency_failures


def main() -> int:
    fixture_master = _find_fixture_master()

    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        exit_code = sfae_main.main(
            ["--input", str(fixture_master), "--output-dir", str(tmp_dir), "--year", "2026"]
        )
        if exit_code != 0:
            print("Generation run FAILED - cannot compare. Check the validation report above.")
            return 1

        generated_files = sorted(tmp_dir.glob("*.xlsx"))
        if not generated_files:
            print("No output workbook was generated.")
            return 1
        generated_path = generated_files[0]

        diffs, consistency_failures = compare_workbooks(generated_path, GOLDEN_SUMMARY, GOLDEN_SHEET)

    ok = True

    if diffs:
        ok = False
        print(f"{len(diffs)} CROSS-FILE DIFFERENCE(S) FOUND vs {GOLDEN_SUMMARY.name}:\n")
        for d in diffs:
            print(
                f"  [{d['label']!s:40.40}] {d['field']:20s} "
                f"golden {d['golden_coord']:>5}={d['expected']!r:>18}  "
                f"generated {d['generated_coord']:>5}={d['actual']!r}"
            )
        print()
    else:
        print(f"Cross-file check: 0 differences vs {GOLDEN_SUMMARY.name}. PASS")

    if consistency_failures:
        ok = False
        print(f"{len(consistency_failures)} QUARTER-MARGIN CONSISTENCY FAILURE(S):\n")
        for d in consistency_failures:
            print(
                f"  [{d['label']!s:40.40}] row {d['row']}: "
                f"sum(Q1..Q4 Margin)={d['quarter_margin_sum']!r}  "
                f"yearly Margin={d['yearly_margin']!r}"
            )
        print()
    else:
        print("Quarter-margin consistency check: every row's Q1..Q4 Margin sums to its yearly Margin. PASS")

    if ok:
        print("\nALL CHECKS PASS")
        return 0

    print(f"\nFAIL - {len(diffs)} cross-file difference(s), {len(consistency_failures)} consistency failure(s).")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
