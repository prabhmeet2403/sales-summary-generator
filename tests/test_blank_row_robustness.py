"""
tests/test_blank_row_robustness.py
=====================================
Verifies that a blank row in the MIDDLE of a section's data block (not
just the normal spacer between sections) has zero effect on the
generated Summary -- every row after it is still read, grouped, and
aggregated exactly as if the blank row weren't there.

This is a property of how the pipeline already works, not a new
mechanism: every row-reading loop in this codebase
(`excel_reader.read_project_rows`, reused for the main sheet, every
prior-year sheet, and indirectly by `monthly_view.build_monthly_sections`;
`comment_mapper.CommentMapper._load` for the comments sheet) scans every
row from the header down to `ws.max_row` unconditionally, skipping a
blank row with `continue` rather than ever `break`-ing out early, and a
row's section membership is decided entirely by its own Sub-Group
(DS-code) value, never by its position or by "still inside a contiguous
block of non-blank rows". This test locks that behavior in rather than
just asserting it exists in the source.

The blank row is produced via direct sheet-XML surgery (clearing one
row's own cells in place -- same row count, nothing else touched)
rather than `openpyxl.load_workbook()` + `ws.insert_rows()` +
`wb.save()`. That round-trip route was tried first and produced a
false positive: this fixture's own Margin column is a live formula
(e.g. `=D5-E5`), and openpyxl -- having no formula engine -- never
carries a cached result forward through any load/save cycle for a
formula cell, including ones the edit never touched. That silently
zeroed out unrelated groups' margins workbook-wide and made them look
incorrectly "blank" -- a confound in the test method, not a real
parsing bug (and, notably, the same underlying openpyxl limitation
this task's Issue 1 fix addresses, just triggered by a different
path). Editing the raw XML in place sidesteps that entirely: every
byte this test doesn't touch, including every other formula's cached
value, stays exactly as the fixture shipped it.

Usage:
    python tests/test_blank_row_robustness.py
"""
from __future__ import annotations

import re
import sys
import tempfile
import zipfile
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from gui.runner import generate_summary  # noqa: E402

FIXTURE_MASTER = Path(__file__).resolve().parent / "fixtures" / "master_2026.xlsx"


def _resolve_sheet_path(workbook_xml: str, rels_xml: str, sheet_title: str) -> str:
    """Same title -> zip-internal-path resolution as
    ``summary_writer._inject_cached_formula_values``, reimplemented
    standalone here so this test has no dependency on that module's
    internals -- only on the fixture actually being a normal OOXML
    package, which is true regardless of what generated it."""
    title_to_rid: dict = {}
    for sheet_tag in re.findall(r"<sheet\b[^>]*/>", workbook_xml):
        name_match = re.search(r'\bname="([^"]*)"', sheet_tag)
        rid_match = re.search(r'\br:id="([^"]*)"', sheet_tag)
        if name_match and rid_match:
            title_to_rid[name_match.group(1)] = rid_match.group(1)

    rid_to_target: dict = {}
    for rel_tag in re.findall(r"<Relationship\b[^>]*/>", rels_xml):
        rid_match = re.search(r'\bId="([^"]*)"', rel_tag)
        target_match = re.search(r'\bTarget="([^"]*)"', rel_tag)
        if rid_match and target_match:
            rid_to_target[rid_match.group(1)] = target_match.group(1)

    rid = title_to_rid[sheet_title]
    target = rid_to_target[rid].lstrip("/")
    return target if target.startswith("xl/") else f"xl/{target}"


def _blank_out_row_in_place(src_path: Path, dest_path: Path, sheet_name: str, row_number: int) -> None:
    """Save a copy of `src_path` with every cell in `row_number` on
    `sheet_name` cleared -- turning an existing data row into a
    genuinely blank one IN PLACE (same row count, same every other
    byte) -- by editing the sheet XML directly rather than round-
    tripping the whole workbook through openpyxl.
    """
    with zipfile.ZipFile(src_path, "r") as zin:
        workbook_xml = zin.read("xl/workbook.xml").decode("utf-8")
        rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        infolist = zin.infolist()
        original_items = {item.filename: zin.read(item.filename) for item in infolist}

    sheet_path = _resolve_sheet_path(workbook_xml, rels_xml, sheet_name)
    xml_text = original_items[sheet_path].decode("utf-8")

    row_match = re.search(rf'<row r="{row_number}"[^>]*>.*?</row>', xml_text, re.DOTALL)
    if not row_match:
        raise AssertionError(f"Row {row_number} not found in {sheet_path}")
    attrs_match = re.match(r'<row r="\d+"([^>]*)>', row_match.group(0))
    blanked_row = f'<row r="{row_number}"{attrs_match.group(1)}></row>'
    xml_text = xml_text[:row_match.start()] + blanked_row + xml_text[row_match.end():]

    modified_items = dict(original_items)
    modified_items[sheet_path] = xml_text.encode("utf-8")
    with zipfile.ZipFile(dest_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in infolist:
            zout.writestr(item, modified_items[item.filename])


def main() -> int:
    problems: list = []

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)

        baseline_out = tmp_path / "baseline"
        baseline_out.mkdir()
        baseline_result = generate_summary(str(FIXTURE_MASTER), str(baseline_out), 2026, progress_cb=lambda m: None)
        if not baseline_result.success:
            print("Baseline generation FAILED - cannot test blank-row robustness.")
            return 1

        # Row 20 ("Databricks Ongoing Support", DS10_Secured) sits
        # between "Church's Chicken & IT Support" (row 19) and
        # "Databricks Planning Project" (row 21), both also
        # DS10_Secured -- blanking it out puts a genuine blank row
        # inside a run of same-section data, not at a section boundary.
        modified_master = tmp_path / "master_with_blank_row.xlsx"
        _blank_out_row_in_place(FIXTURE_MASTER, modified_master, "Sales by Customer- 2026", row_number=20)

        modified_out = tmp_path / "modified"
        modified_out.mkdir()
        modified_result = generate_summary(str(modified_master), str(modified_out), 2026, progress_cb=lambda m: None)

        if not modified_result.success:
            problems.append("Generation FAILED against the fixture with a mid-block row blanked out")
        else:
            baseline_report, modified_report = baseline_result.report, modified_result.report

            # Removing one Databricks row should only affect Databricks'
            # own figures (less revenue/margin from that one row) -- every
            # OTHER group, and the overall group/comment counts, must be
            # completely unaffected by the blank row sitting among them.
            if modified_report.total_groups_processed != baseline_report.total_groups_processed:
                problems.append(
                    f"Groups Processed changed after blanking a mid-block row: "
                    f"baseline={baseline_report.total_groups_processed}, "
                    f"modified={modified_report.total_groups_processed}"
                )
            if modified_report.total_comments_matched != baseline_report.total_comments_matched:
                problems.append(
                    f"Comments Matched changed after blanking a mid-block row: "
                    f"baseline={baseline_report.total_comments_matched}, "
                    f"modified={modified_report.total_comments_matched}"
                )
            if len(modified_report.unmapped_sub_groups) != len(baseline_report.unmapped_sub_groups):
                problems.append("Unmapped Sub-Groups changed after blanking a mid-block row")

            baseline_wb = openpyxl.load_workbook(baseline_result.output_path, data_only=True)
            modified_wb = openpyxl.load_workbook(modified_result.output_path, data_only=True)
            # Only the two AGGREGATED sheets are in scope for this
            # "blank row doesn't break aggregation" comparison. The
            # third sheet ("Sales by Customer- 2026") is a deliberate
            # verbatim copy of the source sheet -- it isn't aggregated
            # at all, so when this test intentionally blanks a row in
            # the SOURCE, that sheet is *supposed* to show the same
            # blank row, exactly like the source does. Comparing it
            # here would be asserting the opposite of what it's for.
            for sheet_name in ("2026", "2026 Actual & Forecast"):
                b_ws, m_ws = baseline_wb[sheet_name], modified_wb[sheet_name]
                if b_ws.max_row != m_ws.max_row:
                    problems.append(f"[{sheet_name}] row count changed: {b_ws.max_row} vs {m_ws.max_row}")
                    continue

                # Compare by (row number, group NAME) rather than raw cell
                # position: a subtotal/total row legitimately changes once
                # Databricks' own aggregate changes (one of its several
                # rows was removed), so only Databricks' own row(s) and any
                # row whose Name contains "subtotal"/"total" are allowed to
                # differ -- every other individual group's row must be
                # completely untouched by a blank row sitting elsewhere in
                # the same section.
                for row in range(1, b_ws.max_row + 1):
                    name_val = b_ws.cell(row=row, column=1).value
                    is_exempt = name_val == "Databricks" or (
                        isinstance(name_val, str) and "subtotal" in name_val.lower()
                    )
                    if is_exempt:
                        continue
                    for col in range(1, b_ws.max_column + 1):
                        b_val = b_ws.cell(row=row, column=col).value
                        m_val = m_ws.cell(row=row, column=col).value
                        if isinstance(b_val, (int, float)) and isinstance(m_val, (int, float)):
                            if abs(b_val - m_val) > 0.005:
                                problems.append(f"[{sheet_name}] cell ({row},{col}), group {name_val!r}: baseline={b_val} modified={m_val}")
                        elif b_val != m_val:
                            problems.append(f"[{sheet_name}] cell ({row},{col}), group {name_val!r}: baseline={b_val!r} modified={m_val!r}")

                # Databricks keeps 3 rows in this fixture; the row we
                # blanked contributed real revenue, so its remaining
                # comment (attached to whichever of its rows carries it)
                # and group membership must still resolve correctly --
                # already covered by Comments Matched / Groups Processed
                # staying identical above, plus every other group's own
                # row(s) being proven byte-for-byte unaffected.

    if problems:
        print("\nFAILURES:")
        for p in problems[:30]:
            print(f"  - {p}")
        if len(problems) > 30:
            print(f"  ... and {len(problems) - 30} more")
        print(f"\nFAIL - {len(problems)} problem(s).")
        return 1

    print("ALL BLANK-ROW ROBUSTNESS CHECKS PASS -- a blank row mid-section only affected that one group, nothing else.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
