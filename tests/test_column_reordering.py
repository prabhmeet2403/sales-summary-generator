"""
tests/test_column_reordering.py
================================
Regression test for Rule 7 ("no hardcoded row numbers, column letters,
or column indexes"): builds a copy of the fixture Master workbook with
every column on every sheet shuffled into a random new order (headers
moving together with their data -- exactly as if someone had dragged
columns around in Excel), runs the real pipeline against both the
original and the scrambled copy, and asserts the two generated Summary
workbooks are cell-for-cell identical.

This test exists because reading the code is not proof: it caught a
real bug (see historical_lookup.py's `_load_fuzzy_sheet` docstring)
that a purely static review of "does this function accept a `path`
argument and call `build_column_map`" would not have found. Re-run this
any time excel_reader.py, aggregator.py, comment_mapper.py, or
historical_lookup.py changes.

Usage:
    python tests/test_column_reordering.py
"""
from __future__ import annotations

import random
import sys
import tempfile
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import main as sfae_main  # noqa: E402

FIXTURE_MASTER = Path(__file__).resolve().parent / "fixtures" / "master_2026.xlsx"
SHEETS_TO_SHUFFLE = [
    "Sales by Customer- 2026",
    "2026_ClientComments",
    "Sales by Customer- 2025",
    "Sales by Customer- 2024",
]
SEEDS_TO_TEST = [42, 999, 7, 123, 2026, 31415, 8675309, 1]


def shuffle_sheet_columns(ws, seed: int) -> None:
    rng = random.Random(seed)
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))

    max_row, max_col = ws.max_row, ws.max_column
    columns = [
        [(ws.cell(r, c).value, ws.cell(r, c).number_format) for r in range(1, max_row + 1)]
        for c in range(1, max_col + 1)
    ]
    order = list(range(max_col))
    rng.shuffle(order)

    for new_c, old_c in enumerate(order, start=1):
        for r, (val, fmt) in enumerate(columns[old_c], start=1):
            cell = ws.cell(r, new_c)
            cell.value = val
            cell.number_format = fmt


def build_shuffled_workbook(seed: int, out_path: Path) -> None:
    wb = openpyxl.load_workbook(FIXTURE_MASTER, data_only=True)
    for sheet_name in SHEETS_TO_SHUFFLE:
        shuffle_sheet_columns(wb[sheet_name], seed)
    wb.save(out_path)


def generate(input_path: Path, output_dir: Path) -> Path:
    exit_code = sfae_main.main(
        ["--input", str(input_path), "--output-dir", str(output_dir), "--year", "2026"]
    )
    if exit_code != 0:
        raise RuntimeError(f"Generation failed for {input_path}")
    files = sorted(output_dir.glob("*.xlsx"))
    if not files:
        raise RuntimeError(f"No output workbook produced for {input_path}")
    return files[0]


def diff_workbooks(path_a: Path, path_b: Path):
    wb_a = openpyxl.load_workbook(path_a, data_only=False)
    wb_b = openpyxl.load_workbook(path_b, data_only=False)
    ws_a = wb_a["2026"]
    ws_b = wb_b["2026"]
    max_row = max(ws_a.max_row, ws_b.max_row)
    max_col = max(ws_a.max_column, ws_b.max_column)
    diffs = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v_a = ws_a.cell(r, c).value
            v_b = ws_b.cell(r, c).value
            if v_a != v_b:
                diffs.append((r, c, v_a, v_b))
    return diffs


def main() -> int:
    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)

        baseline_dir = tmp_dir / "baseline"
        baseline_dir.mkdir()
        baseline_output = generate(FIXTURE_MASTER, baseline_dir)
        print(f"Baseline (unmodified column order) generated: {baseline_output.name}")

        any_failed = False
        for seed in SEEDS_TO_TEST:
            shuffled_path = tmp_dir / f"master_shuffled_{seed}.xlsx"
            build_shuffled_workbook(seed, shuffled_path)

            shuffled_dir = tmp_dir / f"shuffled_{seed}"
            shuffled_dir.mkdir()
            shuffled_output = generate(shuffled_path, shuffled_dir)

            diffs = diff_workbooks(baseline_output, shuffled_output)
            if diffs:
                any_failed = True
                print(f"\nSEED {seed}: {len(diffs)} DIFFERENCE(S) FOUND (column order should not matter!)")
                for r, c, v_a, v_b in diffs:
                    print(f"  Row {r}, Col {c}: baseline={v_a!r}  shuffled={v_b!r}")
            else:
                print(f"SEED {seed}: 0 differences -- column order had no effect. PASS")

    if any_failed:
        print("\nFAIL - output changed when columns were reordered.")
        return 1
    print(f"\nALL {len(SEEDS_TO_TEST)} SHUFFLES MATCH THE BASELINE EXACTLY - PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
