"""
tests/test_future_year_compatibility.py
========================================
Regression test for a real bug found while verifying future-year
compatibility: `aggregator.attach_historical` crashed with
`UnboundLocalError: cannot access local variable 'source'` whenever the
historical lookback reached a year with NO corresponding sheet in the
workbook at all (e.g. asking for target year 2025 when the workbook's
oldest sheet is 2024, so the lookback needs a nonexistent "2023" sheet).

This is guaranteed to happen for the *oldest* year(s) a workbook can be
run against, and would silently block "future-year compatibility" in
reverse -- running the tool against a year near the beginning of a
workbook's history, not just a new one at the end. It was pre-existing
(confirmed by reproducing it against the unmodified project) and
unrelated to the Q1-Q4 Margin/Total column layout; the fix in
aggregator.py only ensures a bookkeeping field always gets a value and
touches no calculated number (that field is write-only -- verified never
read anywhere in the codebase).

This test locks in three things:
1. For years whose own sheet has the modern Group/Sub-Group columns
   (2025, 2026 in the fixture), generation succeeds -- including when
   one or more prior years have no sheet in the workbook at all -- and
   the new grouped Q1-Q4 Margin/Total header survives.
2. For a year whose own sheet predates that schema (2024 in the
   fixture), generation is expected to fail *gracefully* (a controlled,
   reported error) since that sheet cannot serve as a target year's main
   sheet -- never an unhandled exception/crash.
3. Either way, no bare `UnboundLocalError`-style crash escapes main.py.

Usage:
    python tests/test_future_year_compatibility.py
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import main as sfae_main  # noqa: E402

FIXTURE_MASTER = Path(__file__).resolve().parent / "fixtures" / "master_2026.xlsx"

# The fixture workbook's oldest sheet is "Sales by Customer- 2024", so
# running for 2025 requires at least one prior-year sheet (2023) that
# does not exist -- exactly the condition that crashed. 2024 itself is
# included too, but as an EXPECTED graceful failure: that sheet predates
# the Group/Sub-Group column scheme entirely, so it cannot serve as a
# *target* year's main sheet at all (only as a historical fallback for a
# later year) -- main.py is expected to report that clearly, not crash.
YEARS_EXPECTED_TO_SUCCEED = [2025, 2026]
YEARS_EXPECTED_TO_FAIL_GRACEFULLY = [2024]


def _check_year_succeeds(year: int, tmp_dir: Path) -> list:
    problems = []
    out_dir = tmp_dir / str(year)
    out_dir.mkdir()

    try:
        exit_code = sfae_main.main(
            ["--input", str(FIXTURE_MASTER), "--output-dir", str(out_dir), "--year", str(year)]
        )
    except Exception as exc:  # noqa: BLE001
        problems.append(f"year {year}: raised an exception instead of handling it gracefully: {exc!r}")
        return problems

    if exit_code != 0:
        problems.append(f"year {year}: main.py returned exit code {exit_code} (expected success)")
        return problems

    outputs = sorted(out_dir.glob("*.xlsx"))
    if not outputs:
        problems.append(f"year {year}: no output workbook was written")
        return problems

    wb = load_workbook(outputs[0])
    sheet_name = "Multi-Year Revenue & Margin"
    if sheet_name not in wb.sheetnames:
        problems.append(f"year {year}: expected sheet '{sheet_name}' not found (sheets: {wb.sheetnames})")
        return problems

    ws = wb[sheet_name]
    row2 = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    row3 = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
    quarters_found = sum(1 for v in row2 if isinstance(v, str) and v.strip().upper() in ("Q1", "Q2", "Q3", "Q4"))
    if quarters_found != 4:
        problems.append(f"year {year}: expected 4 quarter group headers (Q1-Q4) on row 2, found {quarters_found}")
    if "Margin" not in row3 or "Total" not in row3:
        problems.append(f"year {year}: expected 'Margin'/'Total' sub-labels on row 3, got {row3[:12]}")

    return problems


def _check_year_fails_gracefully(year: int, tmp_dir: Path) -> list:
    """For a year whose own sheet can't serve as a target (e.g. predates
    the Group column), the tool must report a controlled failure
    (non-zero exit, a message in the validation report) -- never an
    unhandled exception/crash."""
    problems = []
    out_dir = tmp_dir / str(year)
    out_dir.mkdir()

    try:
        exit_code = sfae_main.main(
            ["--input", str(FIXTURE_MASTER), "--output-dir", str(out_dir), "--year", str(year)]
        )
    except Exception as exc:  # noqa: BLE001
        problems.append(
            f"year {year}: raised an unhandled exception ({exc!r}) instead of a controlled, "
            "reported failure -- this is the exact crash class this test guards against."
        )
        return problems

    if exit_code == 0:
        problems.append(f"year {year}: expected a graceful failure (no Group column) but generation succeeded")
    return problems


def main() -> int:
    all_problems = []
    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        for year in YEARS_EXPECTED_TO_SUCCEED:
            problems = _check_year_succeeds(year, tmp_dir)
            if problems:
                all_problems.extend(problems)
            else:
                print(f"Year {year}: generated successfully with the grouped quarter layout intact. PASS")
        for year in YEARS_EXPECTED_TO_FAIL_GRACEFULLY:
            problems = _check_year_fails_gracefully(year, tmp_dir)
            if problems:
                all_problems.extend(problems)
            else:
                print(f"Year {year}: failed gracefully (no unhandled exception) as expected. PASS")

    if all_problems:
        print("\nFAILURES:")
        for p in all_problems:
            print(f"  - {p}")
        print(f"\nFAIL - {len(all_problems)} problem(s).")
        return 1

    total = len(YEARS_EXPECTED_TO_SUCCEED) + len(YEARS_EXPECTED_TO_FAIL_GRACEFULLY)
    print(f"\nALL {total} YEARS BEHAVE CORRECTLY - future-year (and past-year) compatibility intact.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
