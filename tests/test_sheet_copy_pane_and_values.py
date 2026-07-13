"""
tests/test_sheet_copy_pane_and_values.py
===========================================
Two focused regression tests for bugs found (and fixed) in
sheet_copy.py after initial delivery:

1. A frozen pane that has ALSO been independently scrolled (its
   `topLeftCell` differs from its actual split point) must round-trip
   exactly -- not through `Worksheet.freeze_panes`'s convenience
   setter, which is lossy for exactly this case (see
   `sheet_copy._copy_sheet_view`'s docstring). Constructed directly
   here (xSplit=3, ySplit=2, topLeftCell="D70") rather than relying on
   any fixture happening to have this same shape, since the bug is
   specifically about frozen-AND-scrolled panes, not simple ones.

2. Every cell in the copied sheet must carry its computed VALUE, never
   a formula -- a formula referencing a sheet this module doesn't also
   copy (e.g. a cross-sheet VLOOKUP) produces both an "external links"
   warning and a #REF!/#N/A when Excel opens the output, which has
   nothing to do with the column-removal feature itself.

Usage:
    python tests/test_sheet_copy_pane_and_values.py
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from sheet_copy import copy_source_sheet_as_new_worksheet  # noqa: E402


def _build_source_workbook(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Source"
    ws["A1"] = "Name"
    ws["B1"] = "Comments"
    ws["C1"] = "Value"
    ws["A2"] = "Row2"
    ws["B2"] = "a comment"
    ws["C2"] = 100

    # A formula referencing a DIFFERENT sheet that this test deliberately
    # does NOT ask sheet_copy to also bring over -- reproducing the
    # exact "VLOOKUP into another tab" real-world scenario.
    other = wb.create_sheet("Other")
    other["A1"] = 42
    ws["D2"] = "=Other!A1*2"

    # A SAME-sheet formula referencing a column AFTER the one about to
    # be removed (C, index 3) -- must stay a live formula, with its own
    # reference shifted left by one once column B (Comments) is gone.
    ws["E2"] = "=C2+1"

    # Frozen pane (columns A-C, rows 1-2 -- xSplit=3, ySplit=2) that has
    # ALSO been scrolled down independently, so topLeftCell is nowhere
    # near the actual split point -- the exact shape that broke the old
    # `freeze_panes = topLeftCell` round-trip.
    ws.freeze_panes = "D3"  # establishes the xSplit=3/ySplit=2 split
    ws.sheet_view.pane.topLeftCell = "D70"  # then scrolled down independently

    wb.save(path)

    # openpyxl has no formula engine, so D2's `=Other!A1*2` has no cached
    # <v> at all after a plain save -- unlike a file real Excel has
    # actually opened and calculated at least once. Inject the value a
    # real Excel session would have cached (42*2=84), reusing the exact
    # same injection helper `SummaryWriter.patch_cached_formula_values`
    # already relies on, rather than a second implementation of it.
    from summary_writer import _inject_cached_formula_values
    _inject_cached_formula_values(path, [("Source", "D2", 84.0), ("Source", "E2", 101.0)])


def main() -> int:
    problems: list = []

    with tempfile.TemporaryDirectory() as tmp:
        source_path = f"{tmp}/source.xlsx"
        _build_source_workbook(source_path)

        out_wb = Workbook()
        out_wb.remove(out_wb.active)
        formula_cache: list = []
        copy_source_sheet_as_new_worksheet(out_wb, source_path, "Source", comments_col=2, formula_cache=formula_cache)
        out_path = f"{tmp}/out.xlsx"
        out_wb.save(out_path)

        new_ws = load_workbook(out_path)["Source"]

        # --- 1. Pane split/scroll fidelity ---
        pane = new_ws.sheet_view.pane
        if pane is None:
            problems.append("Pane is missing entirely on the copied sheet")
        else:
            # Comments (column B, index 2) IS one of the 3 originally-
            # frozen columns (A-C), so removing it correctly leaves only
            # 2 columns frozen (A, C-now-shifted-to-B) -- xSplit must
            # shift too, not stay at 3.
            if pane.xSplit != 2:
                problems.append(f"xSplit should be 2 (one of the 3 frozen columns was removed), got {pane.xSplit}")
            if pane.ySplit != 2:
                problems.append(f"ySplit should be 2 (the TRUE split point, row-based, unaffected by a column removal), got {pane.ySplit} -- this is exactly the old bug (row-1 from topLeftCell instead)")
            # D (column 4) shifts left by one to C once column B is removed.
            if pane.topLeftCell != "C70":
                problems.append(f"topLeftCell should shift from 'D70' to 'C70', got {pane.topLeftCell!r}")
            if pane.state != "frozen":
                problems.append(f"pane.state should be 'frozen', got {pane.state!r}")

        # --- 2. Cross-sheet formula becomes a value ---
        # D2 (source formula "=Other!A1*2") shifts to C2 (Comments/B removed).
        new_cell = new_ws["C2"]
        if isinstance(new_cell.value, str) and new_cell.value.startswith("="):
            problems.append(f"Cell C2 is still a formula ({new_cell.value!r}), should be the computed value 84 (it referenced another sheet)")
        elif new_cell.value != 84:
            problems.append(f"Cell C2 should be the computed value 84, got {new_cell.value!r}")

        # --- 3. Same-sheet formula stays LIVE, with its own reference
        #     shifted for the column removal (E2 "=C2+1" -> D2 "=B2+1",
        #     since Value/C shifted to B). ---
        # E (col 5) shifts to D (col 4) once column B is removed.
        same_sheet_cell = new_ws["D2"]
        if same_sheet_cell.value != "=B2+1":
            problems.append(f"Same-sheet formula should be preserved and shifted to '=B2+1', got {same_sheet_cell.value!r}")

        # --- 4. No CROSS-SHEET formula remains anywhere (same-sheet
        #     formulas, like D2 above, are expected and correct) ---
        cross_sheet_formulas = [
            cell.coordinate for row in new_ws.iter_rows() for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=") and "!" in cell.value
        ]
        if cross_sheet_formulas:
            problems.append(f"Cross-sheet formula(s) found on the copied sheet (should be none): {cross_sheet_formulas}")

        # --- 5. formula_cache correctly tracks the preserved same-sheet
        #     formula's already-computed value (for
        #     SummaryWriter.patch_cached_formula_values to inject later) ---
        cache_entry = next((e for e in formula_cache if e[1] == "D2"), None)
        if cache_entry is None:
            problems.append(f"formula_cache has no entry for the preserved formula at D2; entries: {formula_cache}")
        elif abs(cache_entry[2] - 101.0) > 0.005:  # C2 (=100, now shifted to B2) + 1
            problems.append(f"formula_cache value for D2 should be 101, got {cache_entry[2]}")

        # --- 6. Comments column actually removed ---
        if new_ws.max_column != 4:  # Name, Value, cross-sheet-turned-value, same-sheet-formula
            problems.append(f"Expected 4 columns after removing Comments, got {new_ws.max_column}")
        headers = [new_ws.cell(row=1, column=c).value for c in range(1, new_ws.max_column + 1)]
        if "Comments" in headers:
            problems.append("'Comments' header still present")

        # --- 7. No external references in the output package ---
        import zipfile
        with zipfile.ZipFile(out_path) as zf:
            if any("external" in n.lower() for n in zf.namelist()):
                problems.append("Output package contains an externalLink part")

    if problems:
        print("\nFAILURES:")
        for p in problems:
            print(f"  - {p}")
        print(f"\nFAIL - {len(problems)} problem(s).")
        return 1

    print("ALL SHEET_COPY PANE/VALUE CHECKS PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
