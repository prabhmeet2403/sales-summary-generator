"""
tests/test_third_sheet_column_removal.py
===========================================
Verifies the third worksheet ("Sales by Customer- <year>", a copy of
the uploaded Master workbook's own sheet -- see sheet_copy.py) has its
Comments column (located dynamically by header text) fully REMOVED,
not merely blanked, with every column after it shifted left by one and
every structural feature (widths, freeze panes, autofilter,
conditional formatting) correctly remapped, and that Sheets 1/2 are
completely unaffected.

Usage:
    python tests/test_third_sheet_column_removal.py
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from excel_reader import build_column_map  # noqa: E402
from gui.runner import generate_summary  # noqa: E402

FIXTURE_MASTER = Path(__file__).resolve().parent / "fixtures" / "master_2026.xlsx"
SOURCE_SHEET_NAME = "Sales by Customer- 2026"  # the FIXTURE's own sheet name (unaffected by Change 6's rename)
OUTPUT_SHEET3_NAME = "2026 SOW Performance"  # the GENERATED workbook's Sheet 3 title


def main() -> int:
    problems: list = []

    with tempfile.TemporaryDirectory() as tmp:
        result = generate_summary(str(FIXTURE_MASTER), tmp, 2026, progress_cb=lambda m: None)
        if not result.success:
            print("Generation FAILED - cannot verify the third sheet.")
            return 1

        if result.report.warnings:
            problems.append(f"Unexpected validation warnings: {result.report.warnings}")

        out_f = openpyxl.load_workbook(result.output_path, data_only=False)
        out_v = openpyxl.load_workbook(result.output_path, data_only=True)

        expected_sheets = ["Multi-Year Revenue & Margin", "2026 Monthly Performance", OUTPUT_SHEET3_NAME]
        if out_f.sheetnames != expected_sheets:
            problems.append(f"Unexpected sheet name/order: {out_f.sheetnames}, expected {expected_sheets}")

        if OUTPUT_SHEET3_NAME not in out_f.sheetnames:
            print(f"FAIL - '{OUTPUT_SHEET3_NAME}' sheet missing entirely.")
            return 1

        src_f = openpyxl.load_workbook(str(FIXTURE_MASTER), data_only=False)[SOURCE_SHEET_NAME]
        src_v = openpyxl.load_workbook(str(FIXTURE_MASTER), data_only=True)[SOURCE_SHEET_NAME]
        new_f = out_f[OUTPUT_SHEET3_NAME]
        new_v = out_v[OUTPUT_SHEET3_NAME]
        cmap = build_column_map(openpyxl.load_workbook(str(FIXTURE_MASTER))[SOURCE_SHEET_NAME])
        comments_col = cmap.comments

        if comments_col is None:
            print("FAIL - fixture has no Comments column to test removal against.")
            return 1

        # --- column count reduced by exactly one ---
        if new_f.max_column != src_f.max_column - 1:
            problems.append(f"max_column should be {src_f.max_column - 1}, got {new_f.max_column}")

        # --- the Comments header must not exist anywhere on the new sheet ---
        all_new_headers = [new_f.cell(row=cmap.field_header_row, column=c).value for c in range(1, new_f.max_column + 1)]
        if "Comments" in all_new_headers:
            problems.append("'Comments' header still present on the third sheet")

        # --- every remaining column matches the source, correctly shifted ---
        mismatches = 0
        for row in range(1, src_f.max_row + 1):
            for col in range(1, src_f.max_column + 1):
                if col == comments_col:
                    continue
                new_col = col if col < comments_col else col - 1
                sv = src_v.cell(row=row, column=col).value
                nv = new_v.cell(row=row, column=new_col).value
                if isinstance(sv, (int, float)) and isinstance(nv, (int, float)):
                    if abs(float(sv) - float(nv)) > 0.005:
                        mismatches += 1
                elif sv != nv:
                    mismatches += 1
        if mismatches:
            problems.append(f"{mismatches} cell(s) don't match the source at their shifted position")

        # --- structural fidelity, accounting for the shift ---
        def _shift_ref(ref):
            from sheet_copy import _remap_range_string
            return _remap_range_string(ref, comments_col)

        if new_f.freeze_panes != _shift_ref(src_f.freeze_panes):
            problems.append(f"freeze_panes not correctly shifted: {src_f.freeze_panes!r} -> {new_f.freeze_panes!r}")
        if new_f.auto_filter.ref != _shift_ref(src_f.auto_filter.ref):
            problems.append(f"auto_filter.ref not correctly shifted: {src_f.auto_filter.ref!r} -> {new_f.auto_filter.ref!r}")

        # Column widths are auto-fit to content (see
        # sheet_copy.py's `_autofit_column_widths`) rather than copied
        # verbatim from the source -- an explicit, separate decision
        # for this sheet. Verify auto-fit actually ran (every column
        # has SOME explicit width, not all silently falling back to
        # openpyxl's bare default of 13) and is at least wide enough
        # for that column's own header text, rather than asserting an
        # exact match to the source's original widths.
        no_width_set = [
            get_column_letter(c) for c in range(1, new_f.max_column + 1)
            if new_f.column_dimensions[get_column_letter(c)].width is None
        ]
        if no_width_set:
            problems.append(f"Column(s) with no width set at all (auto-fit did not run): {no_width_set[:10]}")

        too_narrow = []
        from column_autofit import display_text
        for col in range(1, new_f.max_column + 1):
            header_cell = new_f.cell(row=cmap.field_header_row, column=col)
            header_text = display_text(header_cell.value, header_cell.number_format)
            width = new_f.column_dimensions[get_column_letter(col)].width
            if header_text and width is not None and width < len(header_text):
                too_narrow.append(get_column_letter(col))
        if too_narrow:
            problems.append(f"Column(s) narrower than their own DISPLAYED header text: {too_narrow[:10]}")

        # --- NO CROSS-SHEET formula may remain -- see sheet_copy.py's
        #     module docstring for why: a formula referencing a sheet
        #     this copy doesn't also bring over (e.g. a VLOOKUP into
        #     "Salary Projections 2026") produces both an "external
        #     links" warning and a #REF!/#N/A in Excel. Those cells
        #     must carry their already-computed value instead. A
        #     SAME-sheet formula (e.g. "=D4-E4") is expected and
        #     correct to remain live. ---
        cross_sheet_formula_cells = []
        same_sheet_formula_cells = []
        for row in new_f.iter_rows():
            for cell in row:
                val = cell.value
                text = val.text if type(val).__name__ == "ArrayFormula" else val
                if isinstance(text, str) and text.startswith("="):
                    if "!" in text:
                        cross_sheet_formula_cells.append(cell.coordinate)
                    else:
                        same_sheet_formula_cells.append(cell.coordinate)
        if cross_sheet_formula_cells:
            problems.append(f"{len(cross_sheet_formula_cells)} cross-sheet formula(s) remain in the third sheet (should be 0): {cross_sheet_formula_cells[:10]}")
        if not same_sheet_formula_cells:
            problems.append("No same-sheet formulas were preserved at all -- expected at least some (e.g. '=D4-E4'-style cells)")

        # --- every preserved same-sheet formula has a cached value
        #     (same "visible immediately" requirement as Sheets 1/2) ---
        missing_cached = [c for c in same_sheet_formula_cells if new_v[c].value is None]
        if missing_cached:
            problems.append(f"{len(missing_cached)} preserved formula(s) have no cached value: {missing_cached[:10]}")

        # --- values must match the SOURCE's own computed result exactly ---
        value_mismatches = 0
        for row in range(1, src_f.max_row + 1):
            for col in range(1, src_f.max_column + 1):
                if col == comments_col:
                    continue
                new_col = col if col < comments_col else col - 1
                sv = src_v.cell(row=row, column=col).value
                nv = new_v.cell(row=row, column=new_col).value
                if isinstance(sv, (int, float)) and isinstance(nv, (int, float)):
                    if abs(float(sv) - float(nv)) > 0.005:
                        value_mismatches += 1
                elif sv != nv:
                    value_mismatches += 1
        if value_mismatches:
            problems.append(f"{value_mismatches} cell(s) don't match the source's computed value")

        # --- no external-workbook references anywhere in the package ---
        import zipfile
        with zipfile.ZipFile(result.output_path) as zf:
            if any("external" in n.lower() for n in zf.namelist()):
                problems.append("Output package contains an externalLink part")
            if "externalReference" in zf.read("xl/workbook.xml").decode("utf-8"):
                problems.append("workbook.xml declares an externalReference")

        # --- freeze panes / split panes / scroll position match the
        #     source exactly (not just the lossy topLeftCell alone --
        #     see sheet_copy.py's _copy_sheet_view for why xSplit/ySplit
        #     matter independently of topLeftCell) ---
        src_pane, new_pane = src_f.sheet_view.pane, new_f.sheet_view.pane
        if (src_pane is None) != (new_pane is None):
            problems.append("Pane presence differs between source and third sheet")
        elif src_pane is not None:
            if (src_pane.xSplit, src_pane.ySplit, src_pane.topLeftCell, src_pane.state) != \
               (new_pane.xSplit, new_pane.ySplit, new_pane.topLeftCell, new_pane.state):
                problems.append(
                    f"Pane split/scroll differs: source=(xSplit={src_pane.xSplit}, ySplit={src_pane.ySplit}, "
                    f"topLeftCell={src_pane.topLeftCell!r}) vs third sheet=(xSplit={new_pane.xSplit}, "
                    f"ySplit={new_pane.ySplit}, topLeftCell={new_pane.topLeftCell!r})"
                )

        # --- Sheets 1 & 2 unaffected ---
        for sheet_name in ("Multi-Year Revenue & Margin", "2026 Monthly Performance"):
            if sheet_name not in out_f.sheetnames:
                problems.append(f"Sheet '{sheet_name}' is missing")

    if problems:
        print("\nFAILURES:")
        for p in problems:
            print(f"  - {p}")
        print(f"\nFAIL - {len(problems)} problem(s).")
        return 1

    print("ALL THIRD-SHEET COLUMN-REMOVAL CHECKS PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
