"""
tests/test_worksheet1_excludes_projection.py
================================================
Locks in an explicit business rule: the two Projection sections
("Track 1 (Projection)" / "Track 2 (Projection)", Sub-Group codes
DS30_Projection/DS50_Projection -- see
`config.WORKSHEET2_ADDITIONAL_SECTIONS`) belong ONLY on Worksheet 2
("<year> Monthly Performance"), never on Worksheet 1
("Multi-Year Revenue & Margin").

Guards against a regression of the mistake this rule was written to
correct: adding the two DS codes directly to `config.OUTPUT_SECTIONS`
(which both worksheets are ultimately built from) puts the Projection
sections on BOTH worksheets instead of Worksheet 2 alone.

Usage:
    python tests/test_worksheet1_excludes_projection.py
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import config  # noqa: E402
from gui.runner import generate_summary  # noqa: E402

FIXTURE_MASTER = Path(__file__).resolve().parent / "fixtures" / "master_2026.xlsx"

_PROJECTION_KEYS = {"projects_track1_projection", "projects_track2_projection"}
_PROJECTION_TITLES = {"Track 1 (Projection)", "Track 2 (Projection)"}


def main() -> int:
    problems: list = []

    # --- config-level guard: OUTPUT_SECTIONS (Worksheet 1's own list)
    #     must never contain the Projection sections; they belong only
    #     in WORKSHEET2_ADDITIONAL_SECTIONS. ---
    worksheet1_keys = {s.key for s in config.OUTPUT_SECTIONS}
    if worksheet1_keys & _PROJECTION_KEYS:
        problems.append(f"config.OUTPUT_SECTIONS (Worksheet 1) must not include the Projection sections, found: {worksheet1_keys & _PROJECTION_KEYS}")

    worksheet2_extra_keys = {s.key for s in config.WORKSHEET2_ADDITIONAL_SECTIONS}
    if worksheet2_extra_keys != _PROJECTION_KEYS:
        problems.append(f"config.WORKSHEET2_ADDITIONAL_SECTIONS should be exactly the two Projection sections, found: {worksheet2_extra_keys}")

    with tempfile.TemporaryDirectory() as tmp:
        result = generate_summary(str(FIXTURE_MASTER), tmp, 2026, progress_cb=lambda m: None)
        if not result.success:
            print("Generation FAILED - cannot verify Worksheet 1/2 section split.")
            return 1

        import openpyxl
        wb = openpyxl.load_workbook(result.output_path, data_only=True)

        ws1 = wb[wb.sheetnames[0]]
        ws1_titles = {ws1.cell(row=r, column=1).value for r in range(1, ws1.max_row + 1)}
        found_on_ws1 = ws1_titles & _PROJECTION_TITLES
        if found_on_ws1:
            problems.append(f"Worksheet 1 ('{ws1.title}') contains Projection section title(s) that must only be on Worksheet 2: {found_on_ws1}")

        ws2 = wb[wb.sheetnames[1]]
        ws2_titles = {ws2.cell(row=r, column=1).value for r in range(1, ws2.max_row + 1)}
        missing_on_ws2 = _PROJECTION_TITLES - ws2_titles
        if missing_on_ws2:
            problems.append(f"Worksheet 2 ('{ws2.title}') is missing expected Projection section title(s): {missing_on_ws2}")

        # The validation report must still show the Projection sections
        # were processed (for Worksheet 2), and must NOT list their
        # Sub-Group codes as "unmapped" (they ARE handled, just not on
        # Worksheet 1).
        report_text = result.report.render()
        for title in _PROJECTION_TITLES:
            if f"[Section] {title}" not in report_text:
                problems.append(f"Validation report is missing a processed-section entry for '{title}'")
        if "DS30_Projection" in report_text.split("Unmapped Sub-Groups")[-1] if "Unmapped Sub-Groups" in report_text else False:
            problems.append("DS30_Projection incorrectly still listed as unmapped")
        if "DS50_Projection" in report_text.split("Unmapped Sub-Groups")[-1] if "Unmapped Sub-Groups" in report_text else False:
            problems.append("DS50_Projection incorrectly still listed as unmapped")

    if problems:
        print("\nFAILURES:")
        for p in problems:
            print(f"  - {p}")
        print(f"\nFAIL - {len(problems)} problem(s).")
        return 1

    print("ALL WORKSHEET 1 / WORKSHEET 2 PROJECTION-SPLIT CHECKS PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
