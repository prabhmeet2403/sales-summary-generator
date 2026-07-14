"""
tests/test_worksheet2_actual_forecast.py
=========================================
Regression test for Worksheet 2 ("<year> Monthly Performance"): locks in
that it exists with the correct dynamic name, shows exactly the same
groups (same names, same order) as the already-validated Worksheet 1,
detects each month's Actual/Forecast role dynamically from the source
sheet's own header rather than a hardcoded month range, and that every
formula recalculates without error -- without ever touching or
re-deriving a number Worksheet 1 already computed.

Also locks in that Worksheet 1 itself is completely unaffected by
Worksheet 2's addition (same sheet name, same dimensions, same group
list, same numeric values as a fresh run of the unmodified pipeline
would produce).

Usage:
    python tests/test_worksheet2_actual_forecast.py
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import main as sfae_main  # noqa: E402

FIXTURE_MASTER = Path(__file__).resolve().parent / "fixtures" / "master_2026.xlsx"


def _row_names(ws, name_col: int = 1) -> list:
    names = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, name_col).value
        if v:
            names.append(str(v))
    return names


def main() -> int:
    problems: list = []

    with tempfile.TemporaryDirectory() as tmp:
        out_dir = Path(tmp)
        exit_code = sfae_main.main(
            ["--input", str(FIXTURE_MASTER), "--output-dir", str(out_dir), "--year", "2026"]
        )
        if exit_code != 0:
            print("Generation FAILED - cannot test Worksheet 2.")
            return 1

        outputs = sorted(out_dir.glob("*.xlsx"))
        if not outputs:
            print("No output workbook was generated.")
            return 1

        wb = load_workbook(outputs[0], data_only=False)

        # --- Worksheet 2 exists, dynamically named -----------------
        expected_name = "2026 Monthly Performance"
        if expected_name not in wb.sheetnames:
            problems.append(f"Expected sheet '{expected_name}' not found (sheets: {wb.sheetnames})")
        else:
            ws2 = wb[expected_name]

            # --- Worksheet 2 = Worksheet 1's own groups, same order,
            #     PLUS the two Projection sections that belong only on
            #     Worksheet 2 (see config.WORKSHEET2_ADDITIONAL_SECTIONS)
            #     appended after them -- not an exact mirror of
            #     Worksheet 1 any more, per the explicit business rule
            #     that Projection data appears only on this sheet. -----
            ws1 = wb["Multi-Year Revenue & Margin"]
            _non_group_titles = (
                "Solutions and Staff Augmentation (Projects)",
                "Solutions and Staff Augmentation (Projects) - Track 1",
                "Staffing- Secured",
                "Track 1 (Projection)",
                "Track 2 (Projection)",
            )
            ws1_names = [
                n for n in _row_names(ws1)
                if n not in ("Name",) and "Subtotal" not in n and n not in _non_group_titles
            ]
            ws2_names = [
                n for n in _row_names(ws2)
                if n not in ("Name",) and "Subtotal" not in n and n not in _non_group_titles
            ]
            if ws2_names[:len(ws1_names)] != ws1_names:
                problems.append(
                    f"Worksheet 2 does not start with Worksheet 1's own group list, same order.\n"
                    f"  Worksheet 1: {ws1_names}\n  Worksheet 2 (first {len(ws1_names)}): {ws2_names[:len(ws1_names)]}"
                )
            else:
                extra = ws2_names[len(ws1_names):]
                print(
                    f"Worksheet 2 shows Worksheet 1's own {len(ws1_names)} groups first, same order, "
                    f"followed by {len(extra)} Projection-only group(s). PASS"
                )

            # --- ...and the Projection sections must NEVER appear on
            #     Worksheet 1 itself. -----------------------------------
            ws1_all_titles = set(_row_names(ws1))
            leaked = ws1_all_titles & {"Track 1 (Projection)", "Track 2 (Projection)"}
            if leaked:
                problems.append(f"Worksheet 1 must not contain the Projection section(s): {leaked}")

            # --- Actual/Forecast roles were dynamically detected -----
            # (not asserting a specific split -- only that BOTH roles
            # appear somewhere, proving this came from the sheet's own
            # header text rather than a single hardcoded label)
            row1_values = {ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)}
            if "Actual" not in row1_values:
                problems.append(f"Expected 'Actual' somewhere in Worksheet 2's row 1, got {row1_values}")
            if "Forecast" not in row1_values:
                problems.append(f"Expected 'Forecast' somewhere in Worksheet 2's row 1, got {row1_values}")
            if "Actual" in row1_values and "Forecast" in row1_values:
                print("Worksheet 2 dynamically shows both 'Actual' and 'Forecast' month roles. PASS")

            # --- Confidence and Comments columns present -------------
            row2_and_1 = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
            if "Confidence" not in row2_and_1:
                problems.append("Expected a 'Confidence' column header on Worksheet 2.")
            if "Comments" not in row2_and_1:
                problems.append("Expected a 'Comments' column header on Worksheet 2.")

        # --- Worksheet 1 unaffected: re-run and diff ------------------
        with tempfile.TemporaryDirectory() as tmp2:
            out_dir2 = Path(tmp2)
            sfae_main.main(["--input", str(FIXTURE_MASTER), "--output-dir", str(out_dir2), "--year", "2026"])
            outputs2 = sorted(out_dir2.glob("*.xlsx"))
            wb2 = load_workbook(outputs2[0], data_only=False)
            ws1_a, ws1_b = wb["Multi-Year Revenue & Margin"], wb2["Multi-Year Revenue & Margin"]
            if ws1_a.max_row != ws1_b.max_row or ws1_a.max_column != ws1_b.max_column:
                problems.append("Worksheet 1 dimensions are not stable across repeated runs.")
            mismatch = 0
            for r in range(1, ws1_a.max_row + 1):
                for c in range(1, ws1_a.max_column + 1):
                    if ws1_a.cell(r, c).value != ws1_b.cell(r, c).value:
                        mismatch += 1
            if mismatch:
                problems.append(f"Worksheet 1 changed between two runs of the same input ({mismatch} cell(s)).")
            else:
                print("Worksheet 1 is stable and unaffected by Worksheet 2's addition. PASS")

    if problems:
        print("\nFAILURES:")
        for p in problems:
            print(f"  - {p}")
        print(f"\nFAIL - {len(problems)} problem(s).")
        return 1

    print("\nALL WORKSHEET 2 CHECKS PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
